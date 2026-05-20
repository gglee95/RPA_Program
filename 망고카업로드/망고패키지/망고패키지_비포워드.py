"""망고패키지 시트의 매물을 비포워드(BeForward) 에 업로드.

[데이터 소스]
- 망고패키지.py 와 동일한 스프레드시트(GID 1403349305)에서 읽어온다.
- 트리거: AC열(계정) 있음 + AN열(비포워드 업로드일자) 비어있음.
- I열의 "배기량 : xxxcc" 를 직접 사용한다 (포토존 플로우처럼 망고 listing 페이지
  스크랩하지 않음 — 망고패키지 시트는 I열에 배기량이 이미 들어있음).

[업로드 엔진]
- "비포워드 생태계/비포워드_자동화/비포워드_crawling.py" 의 BefowordCrawler 를 사용.
- 이미지 다운로드는 Drive API + 서비스 어카운트 (망고패키지.py 의 방식과 동일).
- 가격은 비포워드 브릿지의 구간별 마크업 로직과 동일.

[열 인덱스 (0-based)]
  D 모델 / E 연식 / F 색상 / G 차량번호 / H VIN / I 구조화 텍스트
  J~S 옵션 / Y 드라이브 / AB 가격 / AC 계정 / AL 실패메모 / AN 비포워드 업로드일자

[비포워드 로그인 정보]
- 환경변수 BEFORWARD_USERNAME / BEFORWARD_PASSWORD 로 오버라이드 가능.
- 미설정 시 비포워드_자동화/config.py 의 기본값 (echam@mangoworldcar.com) 사용.

[실행 예]
  python 망고패키지_비포워드.py                # 모든 pending 행
  python 망고패키지_비포워드.py --row 50       # 50행 한 건
  python 망고패키지_비포워드.py --rows 50-60   # 50~60행
"""
from __future__ import annotations

import argparse
import os
import re
import shutil
import sys
import unicodedata
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_DOWN
from pathlib import Path

# Windows cp949 콘솔에서 em-dash/한글 문자열 출력 깨짐 방지
for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

# ── BefowordCrawler 모듈 경로 추가 ───────────────────────────────────────
# 비포워드_자동화 폴더를 sys.path 에 등록해서 그대로 import 한다.
# (이 폴더에는 config.py 가 없어 포토존 측 같은 swap trick 불필요)
_BF_DIR = (
    Path(__file__).resolve().parent.parent.parent
    / "비포워드 생태계"
    / "비포워드_자동화"
)
if str(_BF_DIR) not in sys.path:
    sys.path.insert(0, str(_BF_DIR))

import 비포워드_crawling
import 엔카_크롤러

BefowordCrawler = 비포워드_crawling.BefowordCrawler
CarInfo         = 엔카_크롤러.CarInfo
OptionItem      = 엔카_크롤러.OptionItem
ENCAR_OPTION_MAP = 엔카_크롤러.OPTION_MAP

import gspread
import requests
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build


def _normalize(s: str) -> str:
    """NFKC 정규화 + 모든 공백/제로폭 제거 (옵션 매칭용)."""
    if not s:
        return ""
    s = unicodedata.normalize("NFKC", s)
    s = re.sub(r"[\s​‌‍﻿ㅤ ]+", "", s)
    return s


_MANGO_TAG_RE = re.compile(r"<[^>]+>")
_MANGO_WS_RE = re.compile(r"\s+")
_MANGO_DISPLACEMENT_RE = re.compile(r"배기량\s*[:\-]?\s*([\d,]+)\s*(?:cc|CC|씨씨)?")
_MANGO_UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/143.0.0.0 Safari/537.36"
)


def _fetch_displacement_from_mango(mango_url: str) -> str:
    """망고카 listing 페이지에서 배기량(숫자만) 추출.

    XPath 위치보다 라벨 기반 매칭이 페이지 구조 변경에 강하다.
    실패 시 즉시 ValueError (글로벌 규칙: 폴백 금지).
    """
    if not mango_url:
        raise ValueError("mango_url 이 비어있어 배기량 조회 불가")
    resp = requests.get(
        mango_url,
        headers={"User-Agent": _MANGO_UA, "Accept-Language": "ko-KR,ko;q=0.9"},
        timeout=20,
    )
    resp.raise_for_status()
    text = _MANGO_WS_RE.sub(" ", _MANGO_TAG_RE.sub(" ", resp.text))
    m = _MANGO_DISPLACEMENT_RE.search(text)
    if not m:
        raise ValueError(f"페이지에서 '배기량' 값을 찾지 못함 (url={mango_url[:120]})")
    return m.group(1).replace(",", "")


def _build_upload_price(base_price: str) -> str:
    """비포워드 브릿지 _build_upload_price 와 동일한 구간별 마크업 계산."""
    if not base_price:
        return ""
    try:
        base_value = Decimal(str(base_price))
    except InvalidOperation:
        return ""

    if   base_value <= Decimal("1000"):  markup = Decimal("263")
    elif base_value <= Decimal("1500"):  markup = Decimal("278")
    elif base_value <= Decimal("2000"):  markup = Decimal("283")
    elif base_value <= Decimal("3000"):  markup = Decimal("303")
    elif base_value <= Decimal("5000"):  markup = Decimal("358")
    elif base_value <= Decimal("6000"):  markup = Decimal("388")
    elif base_value <= Decimal("7000"):  markup = Decimal("410")
    elif base_value <= Decimal("8000"):  markup = Decimal("439")
    elif base_value <= Decimal("10000"): markup = Decimal("495")
    elif base_value <= Decimal("15000"): markup = Decimal("630")
    elif base_value <= Decimal("20000"): markup = Decimal("739")
    else:
        markup = (base_value * Decimal("0.05")).quantize(
            Decimal("1"), rounding=ROUND_DOWN
        )

    return str(int((base_value + markup).quantize(Decimal("1"), rounding=ROUND_DOWN)))


class MangoPackageBeforwardUploader:
    # ── 스프레드시트 (망고패키지.py 와 동일) ──────────────────────────
    SPREADSHEET_ID = "1yHN0UM8Rr_CmMjz5fI3CEdhQjHM7VQIaqitWPRIGR8E"
    SHEET_GID = 1403349305
    SERVICE_ACCOUNT_FILE = os.path.join(
        os.path.dirname(__file__),
        "..", "망고카 오토", "adjustmentdata-51a7199ac3ba.json",
    )
    SCOPES = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive.readonly",
    ]

    # ── 열 인덱스 (0-based) ──────────────────────────────────────────
    COL_MODEL          = 3   # D
    COL_YEAR           = 4   # E
    COL_COLOR          = 5   # F
    COL_CAR_NUMBER     = 6   # G
    COL_VIN            = 7   # H
    COL_DETAIL         = 8   # I  (구조화 텍스트 - 배기량 포함)
    COL_OPT_START      = 9   # J
    COL_OPT_END        = 18  # S
    COL_DRIVE_LINK     = 24  # Y
    COL_PLATFORM_URL   = 25  # Z   (망고카 listing URL — 배기량 스크랩 소스)
    COL_PRICE          = 27  # AB
    COL_ACCOUNT        = 28  # AC  (트리거 - 계정 있어야 처리)
    COL_BF_FAIL_NOTE   = 38  # AM  (비포워드 업로드 에러 메모)
    COL_BF_UPLOAD_DATE = 39  # AN  (비포워드 업로드일자)
    COL_BF_LINK        = 40  # AO  (비포워드 매물 링크)

    DOWNLOAD_FOLDER = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "downloaded_images",
    )

    def __init__(self):
        self.worksheet = None
        self.all_rows: list[list[str]] = []
        self.creds = None
        self.drive_links: dict[int, str] = {}
        self._vehicle_ref: list[dict] = []   # 재원표 캐시 (영문→한글 매핑용)

    # ── 스프레드시트 ────────────────────────────────────────────────
    def setup_spreadsheet(self) -> bool:
        try:
            print("[진행] 스프레드시트 연결 중...")
            self.creds = Credentials.from_service_account_file(
                self.SERVICE_ACCOUNT_FILE, scopes=self.SCOPES
            )
            gc = gspread.authorize(self.creds)
            spreadsheet = gc.open_by_key(self.SPREADSHEET_ID)
            self.worksheet = next(
                (s for s in spreadsheet.worksheets() if s.id == self.SHEET_GID),
                None,
            )
            if not self.worksheet:
                print(f"[오류] GID {self.SHEET_GID} 시트를 찾을 수 없습니다")
                return False
            print(f"[OK] 시트 연결: {self.worksheet.title}")
            self.all_rows = self.worksheet.get_all_values()
            self._fetch_drive_links()
            return True
        except Exception as e:
            print(f"[오류] 스프레드시트 연결 실패: {e}")
            return False

    def _fetch_drive_links(self) -> None:
        """Y열 하이퍼링크 일괄 조회 (Sheets API v4)."""
        try:
            service = build("sheets", "v4", credentials=self.creds)
            result = service.spreadsheets().get(
                spreadsheetId=self.SPREADSHEET_ID,
                ranges=[f"'{self.worksheet.title}'!Y:Y"],
                fields="sheets(data(rowData(values(hyperlink,formattedValue,userEnteredValue))))",
            ).execute()
            row_data = (
                result.get("sheets", [{}])[0]
                .get("data", [{}])[0]
                .get("rowData", [])
            )
            for idx, row in enumerate(row_data):
                row_num = idx + 1
                if row_num < 2:
                    continue
                values = row.get("values", [])
                if not values:
                    continue
                cell = values[0]
                link = cell.get("hyperlink", "")
                if not link:
                    uev = cell.get("userEnteredValue", {})
                    formula = uev.get("formulaValue", "") if isinstance(uev, dict) else ""
                    if formula:
                        m = re.search(r'HYPERLINK\("([^"]+)"', formula)
                        if m:
                            link = m.group(1)
                if not link:
                    fv = cell.get("formattedValue", "")
                    if fv and fv.startswith("http"):
                        link = fv
                if link and ("drive.google.com" in link or "docs.google.com" in link):
                    self.drive_links[row_num] = link
            print(f"[OK] Y열 드라이브 링크 {len(self.drive_links)}개 확인")
        except Exception as e:
            print(f"[경고] Y열 일괄 조회 실패: {e}")

    def _get_drive_link(self, row_idx: int) -> str:
        if row_idx in self.drive_links:
            return self.drive_links[row_idx]
        idx = row_idx - 1
        row = self.all_rows[idx] if 0 <= idx < len(self.all_rows) else []
        val = row[self.COL_DRIVE_LINK].strip() if len(row) > self.COL_DRIVE_LINK else ""
        if val.startswith("http") and "drive.google.com" in val:
            return val
        return ""

    def get_pending_rows(self) -> list[dict]:
        """AC 계정 있음 + AN 비포워드일자 비어있음."""
        pending = []
        for row_idx, row in enumerate(self.all_rows[1:], start=2):
            ac = row[self.COL_ACCOUNT] if len(row) > self.COL_ACCOUNT else ""
            an = row[self.COL_BF_UPLOAD_DATE] if len(row) > self.COL_BF_UPLOAD_DATE else ""
            if not ac.strip() or an.strip():
                continue
            pending.append({"row_idx": row_idx, "row": row})
        return pending

    # ── I열 파싱 ───────────────────────────────────────────────────
    @staticmethod
    def parse_i_column(i_val: str) -> dict[str, str]:
        result = {
            "sub_model": "", "drive_type": "", "transmission": "", "fuel": "",
            "seating": "", "mileage": "", "handle": "", "engine_displacement": "",
        }
        if not i_val:
            return result
        patterns = {
            "sub_model":           r"1\.\s*세부모델[ \t]*:[ \t]*([^\r\n]*)",
            "drive_type":          r"2\.\s*구동방식[ \t]*:[ \t]*([^\r\n]*)",
            "transmission":        r"3\.\s*변속기[ \t]*:[ \t]*([^\r\n]*)",
            "fuel":                r"4\.\s*연료[ \t]*:[ \t]*([^\r\n]*)",
            "seating":             r"5\.\s*승차인원[ \t]*:[ \t]*([^\r\n]*)",
            "mileage":             r"6\.\s*주행거리[ \t]*:[ \t]*([^\r\n]*)",
            "handle":              r"7\.\s*핸들위치[ \t]*:[ \t]*([^\r\n]*)",
            "engine_displacement": r"배기량[ \t]*:[ \t]*([^\r\n]*)",
        }
        for key, pattern in patterns.items():
            m = re.search(pattern, i_val)
            if m:
                result[key] = m.group(1).strip()
        return result

    @classmethod
    def _row_options(cls, row: list[str]) -> list[str]:
        names: list[str] = []
        for col_idx in range(cls.COL_OPT_START, cls.COL_OPT_END + 1):
            if col_idx < len(row):
                val = row[col_idx].strip()
                if val:
                    names.append(val)
        return names

    # ── 재원표 영문→한글 keyword 매핑 ─────────────────────────────
    def _load_vehicle_ref(self) -> None:
        """비포워드_자동화 폴더의 국산차/수입차 재원표를 메모리에 로드.

        BefowordCrawler 내부의 _vehicle_ref_table 과 동일한 데이터지만,
        crawler 인스턴스 만들기 전에 우리 쪽에서 lookup 해야 해서 별도 로드.
        """
        if self._vehicle_ref:
            return
        import openpyxl
        for fname in ("국산차 재원표.xlsx", "수입차 재원표.xlsx"):
            fpath = _BF_DIR / fname
            if not fpath.exists():
                continue
            try:
                wb = openpyxl.load_workbook(fpath, data_only=True)
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        keyword  = row[1] if len(row) > 1 else None
                        bf_model = row[3] if len(row) > 3 else None
                        if not keyword and not bf_model:
                            continue
                        self._vehicle_ref.append({
                            "keyword":  str(keyword).strip()  if keyword  else "",
                            "bf_model": str(bf_model).strip() if bf_model else "",
                        })
            except Exception as e:
                print(f"  [경고] 재원표 로드 실패 ({fname}): {e}")

    def _resolve_model_keyword(self, raw_model: str) -> str:
        """망고패키지 D열 영문 모델명을 재원표의 한글 keyword 로 변환.

        매칭 우선순위:
          1) keyword(한글) 단어 경계 매칭
          2) bf_model(영문, 대소문자 무시) 단어 경계 매칭
        매칭 실패 시 원본 그대로 반환 (BefowordCrawler 가 다음 단계에서 명확한
        '재원표 미등록 차종' 에러를 던지도록).
        """
        if not raw_model:
            return raw_model
        self._load_vehicle_ref()
        if not self._vehicle_ref:
            return raw_model

        def _word_match(needle: str, haystack: str, ignore_case: bool = False) -> bool:
            if not needle:
                return False
            pattern = r"(?<![가-힣a-zA-Z0-9])" + re.escape(needle) + r"(?![가-힣a-zA-Z0-9])"
            flags = re.IGNORECASE if ignore_case else 0
            return bool(re.search(pattern, haystack, flags))

        matches: list[str] = []
        for entry in self._vehicle_ref:
            kw = entry["keyword"]
            bm = entry["bf_model"]
            if _word_match(kw, raw_model):
                matches.append(kw)
            elif _word_match(bm, raw_model, ignore_case=True):
                matches.append(kw)
        if not matches:
            return raw_model
        return max(matches, key=len)   # 가장 구체적인 매칭

    @staticmethod
    def _map_options(option_names: list[str]) -> list[OptionItem]:
        """망고패키지 J~S 옵션명 → 엔카_크롤러.OPTION_MAP 경유 BeForward 라벨 매핑.

        매핑 불가 옵션은 조용히 스킵 (글로벌 규칙: 임의 매핑/폴백 금지).
        """
        items: list[OptionItem] = []
        seen: set[str] = set()
        for name in option_names:
            mapped = ENCAR_OPTION_MAP.get(name) or ENCAR_OPTION_MAP.get(_normalize(name))
            if mapped and mapped not in seen:
                items.append(OptionItem(name=name, mapped_name=mapped))
                seen.add(mapped)
        return items

    # ── CarInfo 빌드 ───────────────────────────────────────────────
    def build_car_info(
        self, row: list[str], row_idx: int, drive_link: str
    ) -> CarInfo:
        get = lambda i: (row[i].strip() if len(row) > i else "")

        model       = get(self.COL_MODEL)
        year        = get(self.COL_YEAR)
        color       = get(self.COL_COLOR)
        vin         = get(self.COL_VIN)
        price_raw   = get(self.COL_PRICE)
        platform_url = get(self.COL_PLATFORM_URL)
        i_val       = row[self.COL_DETAIL] if len(row) > self.COL_DETAIL else ""
        opt_names   = self._row_options(row)

        detail = self.parse_i_column(i_val)
        sub_model    = detail["sub_model"]
        mileage_raw  = detail["mileage"]
        transmission = detail["transmission"]
        fuel         = detail["fuel"]
        seating      = detail["seating"]

        # 필수 컬럼 검증 (글로벌 규칙: 폴백/기본값 금지 → 빈 값은 즉시 실패)
        # 배기량은 Z열 망고카 페이지에서 스크랩하므로 Z열만 필수로 체크
        missing = []
        if not model:        missing.append("D열(모델)")
        if not year:         missing.append("E열(연식)")
        if not color:        missing.append("F열(색상)")
        if not vin:          missing.append("H열(VIN)")
        if not mileage_raw:  missing.append("I열 주행거리")
        if not transmission: missing.append("I열 변속기")
        if not fuel:         missing.append("I열 연료")
        if not seating:      missing.append("I열 승차인원")
        if not price_raw:    missing.append("AB열(가격)")
        if not platform_url: missing.append("Z열(망고카 링크)")
        if not drive_link:   missing.append("Y열(드라이브)")
        if missing:
            raise RuntimeError(
                f"행 {row_idx}: 필수 컬럼 비어있음 — {', '.join(missing)}"
            )

        # 배기량: Z열 망고카 페이지에서 스크랩 (실패 시 ValueError → 행 FAIL)
        displacement_digits = _fetch_displacement_from_mango(platform_url)
        print(f"  [배기량] Z열 망고 페이지에서 {displacement_digits} cc 추출")

        # sub_model이 placeholder("-")만 들어있으면 차종에 붙이지 않는다.
        # "TIGUAN -" 같은 결합이 재원표 단어 경계 매칭을 깨므로.
        if sub_model and re.fullmatch(r"[\-_\s]+", sub_model):
            sub_model = ""

        # 영문 모델명 → 재원표의 한글 keyword 로 변환.
        # BefowordCrawler._lookup_vehicle_ref 는 keyword(한글) 기준 단어 경계
        # 매칭이라 영문 입력("TIGUAN")으로는 매칭 실패. 우리 쪽에서 미리 변환.
        model_resolved = self._resolve_model_keyword(model)
        if model_resolved != model:
            print(f"  [모델 매핑] D열 {model!r} → 재원표 keyword {model_resolved!r}")
        model = model_resolved

        car_type = f"{model} {sub_model}".strip() if sub_model else model

        # 숫자만 추출 — 빈 값이면 즉시 실패
        price_digits = "".join(c for c in price_raw if c.isdigit())
        if not price_digits:
            raise RuntimeError(f"행 {row_idx}: AB열 가격에서 숫자 추출 실패 (값={price_raw!r})")
        upload_price = _build_upload_price(price_digits)
        if not upload_price:
            raise RuntimeError(f"행 {row_idx}: 마크업 가격 계산 실패 (원본={price_digits})")

        mileage_digits = "".join(c for c in mileage_raw if c.isdigit())
        if not mileage_digits:
            raise RuntimeError(f"행 {row_idx}: I열 주행거리에서 숫자 추출 실패 (값={mileage_raw!r})")

        options = self._map_options(opt_names)

        car_info = CarInfo(
            car_type=car_type,
            year_month=year,
            mileage=mileage_digits,
            displacement=displacement_digits,
            fuel_type=fuel,
            transmission=transmission,
            color=color,
            seating_capacity=seating,
            price=upload_price,
            inspection_chassis_no=vin,
            options=options,
        )
        car_info.drive_link = drive_link
        car_info.sheet_row = row_idx
        return car_info

    # ── 이미지 다운로드 (requests + embeddedfolderview — 포토존 방식) ──
    def _extract_folder_id(self, drive_link: str) -> str:
        m = re.search(r"/folders/([a-zA-Z0-9_-]+)", drive_link)
        return m.group(1) if m else ""

    _IMG_EXTS_FOR_DL = {".jpg", ".jpeg", ".png", ".gif", ".bmp", ".webp", ".heic", ".heif"}
    _MIN_VALID_BYTES = 200
    _IMAGE_MAGIC_PREFIXES = (
        b"\xFF\xD8\xFF",          # JPG
        b"\x89PNG\r\n\x1a\n",     # PNG
        b"GIF87a", b"GIF89a",     # GIF
        b"RIFF",                  # WebP
        b"BM",                    # BMP
    )
    _FLIP_OPEN_RE = re.compile(r'<div[^>]*class="flip-entry"')

    @classmethod
    def _is_valid_image(cls, content: bytes) -> bool:
        if len(content) < cls._MIN_VALID_BYTES:
            return False
        return any(content[:16].startswith(magic) for magic in cls._IMAGE_MAGIC_PREFIXES)

    @classmethod
    def _list_drive_folder(cls, folder_id: str, session) -> list[dict]:
        """embeddedfolderview HTML → [{id, name, kind}] 목록 (포토존 측 로직과 동일)."""
        url = f"https://drive.google.com/embeddedfolderview?id={folder_id}#list"
        resp = session.get(url, timeout=30)
        resp.raise_for_status()
        html = resp.text
        starts = [m.start() for m in cls._FLIP_OPEN_RE.finditer(html)]
        if not starts:
            return []
        starts.append(len(html))
        entries: list[dict] = []
        for i in range(len(starts) - 1):
            block = html[starts[i]:starts[i + 1]]
            href_m = re.search(r'href="([^"]+)"', block)
            title_m = re.search(r'flip-entry-title[^>]*>([^<]+)<', block)
            if not href_m:
                continue
            href = href_m.group(1)
            name = title_m.group(1).strip() if title_m else ""
            kind, fid = None, None
            m = re.search(r"/file/d/([a-zA-Z0-9_-]+)", href)
            if m:
                kind, fid = "file", m.group(1)
            else:
                m = re.search(r"/drive/folders/([a-zA-Z0-9_-]+)", href)
                if m:
                    kind, fid = "folder", m.group(1)
            if fid:
                entries.append({"id": fid, "name": name, "kind": kind})
        return entries

    @classmethod
    def _collect_image_files(cls, folder_id: str, session, prefix: str = "") -> list[dict]:
        """폴더를 재귀적으로 순회하며 이미지 파일을 수집.

        Returns: [{"id", "name", "sort_key"}] — sort_key 는 폴더 경로 + 자연 정렬 키.
        """
        entries = cls._list_drive_folder(folder_id, session)
        # 같은 폴더 안에서는 파일명 자연 정렬
        def _nkey(s: str):
            parts = re.split(r"(\d+)", os.path.splitext(s)[0])
            return [int(p) if p.isdigit() else p.lower() for p in parts]
        entries.sort(key=lambda e: (e["kind"] != "folder", _nkey(e["name"])))

        out: list[dict] = []
        for e in entries:
            ext = os.path.splitext(e["name"])[1].lower()
            if e["kind"] == "file" and ext in cls._IMG_EXTS_FOR_DL:
                out.append({
                    "id": e["id"],
                    "name": e["name"],
                    "sort_key": f"{prefix}/{e['name']}",
                })
            elif e["kind"] == "folder":
                sub_prefix = f"{prefix}/{e['name']}" if prefix else e["name"]
                out.extend(cls._collect_image_files(e["id"], session, sub_prefix))
        return out

    @classmethod
    def _download_one(cls, file_id: str, dest: str, session) -> bool:
        url = (
            "https://drive.usercontent.google.com/download"
            f"?id={file_id}&export=download&confirm=t&authuser=0"
        )
        try:
            resp = session.get(url, timeout=120)
            resp.raise_for_status()
            content = resp.content
            if not cls._is_valid_image(content):
                return False
            with open(dest, "wb") as fh:
                fh.write(content)
            return True
        except Exception:
            return False

    def download_images_for_row(self, drive_link: str, row_idx: int) -> list[str]:
        """BefowordCrawler._download_images_from_drive_link 의 monkey-patch 대상.

        requests + embeddedfolderview 로 폴더 재귀 순회 → 이미지 다운로드.
        포토존 bf_drive_downloader 와 동일한 방식 (Drive API 미사용, 폴더가
        'anyone with link' 공유여야 함).
        """
        if not drive_link:
            return []
        folder_id = self._extract_folder_id(drive_link)
        if not folder_id:
            print(f"  [경고] 폴더 ID 추출 실패: {drive_link[:60]}")
            return []

        row_folder = os.path.join(self.DOWNLOAD_FOLDER, f"row_{row_idx}")
        if os.path.exists(row_folder):
            shutil.rmtree(row_folder, ignore_errors=True)
        os.makedirs(row_folder, exist_ok=True)

        session = requests.Session()
        session.headers.update({"User-Agent": _MANGO_UA})

        try:
            img_files = self._collect_image_files(folder_id, session)
        except Exception as e:
            print(f"  [경고] 드라이브 폴더 목록 조회 실패: {e}")
            return []
        if not img_files:
            print(f"  [경고] 드라이브 폴더에 이미지 없음 (또는 공유 권한 문제)")
            return []
        print(f"  [Drive] 이미지 {len(img_files)}장 다운로드 시작")

        # 이름 충돌 방지 — 같은 이름 파일이 여러 서브폴더에 있을 수 있으니 인덱스 prefix.
        downloaded: list[str] = []
        for idx, fi in enumerate(img_files, start=1):
            safe_name = f"{idx:03d}_{fi['name']}"
            dst = os.path.join(row_folder, safe_name)
            ok = self._download_one(fi["id"], dst, session)
            if not ok:
                # 1회 재시도
                if os.path.exists(dst):
                    try:
                        os.remove(dst)
                    except Exception:
                        pass
                ok = self._download_one(fi["id"], dst, session)
            if ok:
                downloaded.append(dst)
            else:
                print(f"  [경고] {fi['name']} 다운로드 최종 실패")
        print(f"  [Drive] {len(downloaded)}/{len(img_files)}장 완료")
        return downloaded

    # ── 시트 기록 ──────────────────────────────────────────────────
    @staticmethod
    def _col_letter(col_idx_zero_based: int) -> str:
        col = col_idx_zero_based + 1
        s = ""
        while col > 0:
            col, rem = divmod(col - 1, 26)
            s = chr(65 + rem) + s
        return s

    def mark_failed(self, row_idx: int, reason: str) -> None:
        try:
            letter = self._col_letter(self.COL_BF_FAIL_NOTE)
            self.worksheet.update_acell(f"{letter}{row_idx}", reason[:300])
            print(f"  [시트] {letter}{row_idx} = {reason[:80]!r}")
        except Exception as e:
            print(f"  [경고] 실패메모 기록 실패: {e}")

    def mark_success(self, row_idx: int, bf_url: str) -> None:
        today = datetime.now().strftime("%Y-%m-%d")
        date_letter = self._col_letter(self.COL_BF_UPLOAD_DATE)
        link_letter = self._col_letter(self.COL_BF_LINK)
        try:
            self.worksheet.update_acell(f"{date_letter}{row_idx}", today)
            print(f"  [시트] {date_letter}{row_idx} = {today}")
        except Exception as e:
            print(f"  [경고] {date_letter}{row_idx} 업로드일자 기록 실패: {e}")
        try:
            self.worksheet.update_acell(f"{link_letter}{row_idx}", bf_url)
            print(f"  [시트] {link_letter}{row_idx} = {bf_url}")
        except Exception as e:
            print(f"  [경고] {link_letter}{row_idx} 링크 기록 실패: {e}")

    # ── 업로드 ────────────────────────────────────────────────────
    def upload_row(self, row: list[str], row_idx: int) -> tuple[bool, str]:
        """단일 행 비포워드 업로드. (success, detail) 반환."""
        try:
            drive_link = self._get_drive_link(row_idx)
            car_info = self.build_car_info(row, row_idx, drive_link)
        except Exception as e:
            return False, f"데이터 빌드 실패 — {e}"

        crawler = BefowordCrawler(headless=False)
        # 이미지 다운로드를 Drive API 기반 함수로 교체
        crawler._download_images_from_drive_link = self.download_images_for_row

        try:
            if not crawler.login():
                return False, "비포워드 로그인 실패"
            ok = crawler.fill_vehicle_data(car_info, auto_submit=True)
            submitted = getattr(crawler, "_listing_submitted", False)
            if ok or submitted:
                listing_id = getattr(car_info, "_listing_id", "")
                if listing_id:
                    url = f"https://external-vendor.beforward.jp/tempVehDetails/edit/{listing_id}"
                    return True, url
                return True, "ok (listing ID 미확인)"
            step  = getattr(crawler, "_last_error_step", "") or "unknown"
            cause = getattr(crawler, "_last_error_cause", "") or "fill_vehicle_data returned False"
            return False, f"{step}: {cause}"
        except Exception as e:
            return False, f"예외 — {type(e).__name__}: {str(e)[:200]}"
        finally:
            try:
                crawler.close()
            except Exception:
                pass

    # ── 메인 루프 ─────────────────────────────────────────────────
    def process_all(
        self,
        start_row: int | None = None,
        end_row: int | None = None,
        row_subset: set[int] | None = None,
        force: bool = False,
    ) -> None:
        if force and row_subset:
            # --force + --row/--rows: pending 필터(AN 비어있음) 우회.
            # 명시한 행을 그대로 처리 (재시도 / 디버그용).
            pending = []
            for r in row_subset:
                idx = r - 1
                if 0 <= idx < len(self.all_rows):
                    pending.append({"row_idx": r, "row": self.all_rows[idx]})
        else:
            pending = self.get_pending_rows()
            if start_row is not None:
                pending = [p for p in pending if p["row_idx"] >= start_row]
            if end_row is not None:
                pending = [p for p in pending if p["row_idx"] <= end_row]
            if row_subset is not None:
                pending = [p for p in pending if p["row_idx"] in row_subset]

        if not pending:
            print("[알림] 비포워드 업로드 대상 행이 없습니다.")
            return

        print(f"\n[진행] 비포워드 업로드 대상 {len(pending)}건: "
              f"{[p['row_idx'] for p in pending]}")
        ok_count = 0
        fail_count = 0
        for item in pending:
            row_idx = item["row_idx"]
            row = item["row"]
            print(f"\n{'─'*60}\n[{row_idx}행] 비포워드 업로드 시작")
            success, detail = self.upload_row(row, row_idx)
            if success:
                self.mark_success(row_idx, detail)
                print(f"  [OK] {detail}")
                ok_count += 1
            else:
                self.mark_failed(row_idx, detail)
                print(f"  [FAIL] {detail}")
                fail_count += 1
        print(f"\n{'='*60}\n[요약] 성공 {ok_count} / 실패 {fail_count}")


def _parse_row_spec(spec: str) -> set[int]:
    wanted: set[int] = set()
    for part in spec.split(","):
        part = part.strip()
        if not part:
            continue
        if "-" in part:
            lo, hi = part.split("-", 1)
            wanted.update(range(int(lo), int(hi) + 1))
        else:
            wanted.add(int(part))
    return wanted


def main() -> int:
    parser = argparse.ArgumentParser(description="망고패키지 시트 → 비포워드 업로드")
    g = parser.add_mutually_exclusive_group()
    g.add_argument("--row", type=int, help="단일 행")
    g.add_argument("--rows", type=str, help="행 범위/콤마 (예: 50-60 또는 50,52,55)")
    g.add_argument("--start", type=int, help="시작 행 (--end 와 같이 사용)")
    parser.add_argument("--end", type=int, help="끝 행 (--start 와 같이 사용)")
    parser.add_argument(
        "--force",
        action="store_true",
        help="--row/--rows 와 함께 사용 시 pending 필터(AN 비어있음) 우회. 재시도용.",
    )
    args = parser.parse_args()

    uploader = MangoPackageBeforwardUploader()
    if not uploader.setup_spreadsheet():
        return 1

    row_subset: set[int] | None = None
    start_row: int | None = None
    end_row: int | None = None
    if args.row:
        row_subset = {args.row}
    elif args.rows:
        row_subset = _parse_row_spec(args.rows)
    else:
        start_row = args.start
        end_row = args.end

    uploader.process_all(
        start_row=start_row, end_row=end_row, row_subset=row_subset, force=args.force,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
