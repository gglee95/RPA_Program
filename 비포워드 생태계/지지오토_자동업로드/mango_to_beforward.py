"""
망고월드카 지지오토 차량 → 비포워드 자동 업로드
1. Google Sheets(망고카 통합 시트)에서 지지오토 API 게시 차량 수집
2. CarInfo 객체로 변환
3. BefowordCrawler.fill_vehicle_data() 로 비포워드 업로드
4. 판매 완료 차량 감지 → BeforwardSuspensionManager 로 게시 정지
"""
import sys
import os
import re
import time
import logging
import traceback
from pathlib import Path
from datetime import datetime

HERE = Path(__file__).parent
if str(HERE) not in sys.path:
    sys.path.insert(0, str(HERE))

from 비포워드_crawling import BefowordCrawler           # noqa: E402
from 엔카_크롤러 import CarInfo, OptionItem             # noqa: E402
from beforward_suspension_manager import BeforwardSuspensionManager  # noqa: E402

LOG_FILE = HERE / "mango_to_beforward.log"
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(),
    ],
)
log = logging.getLogger(__name__)

# ── 설정 ─────────────────────────────────────────────────────────────────────
DETAIL_BASE = "https://mangoworldcar.com/ko/car-detail"
OPTION_MAPPING_FILE = HERE / "비포워드 엔카 옵션_망고카 추가.xlsx"
OPTION_MAPPING_SHEET = "망고카"

# 비포워드 계정 (비포워드_자동화/config.py 에서 읽어오거나 직접 지정)
os.environ.setdefault("BEFORWARD_USERNAME", "echam@mangoworldcar.com")
os.environ.setdefault("BEFORWARD_PASSWORD", "VJSXaPQR")

# 테스트 시 True → 자동 제출 안 함 (폼 채운 상태로 멈춤)
DRY_RUN = False

# ── Google Sheets 설정 ────────────────────────────────────────────────────────
# 망고카 통합 데이터 시트 (gid=1710203054)
MANGO_SHEET_ID = "1P6AJOgbyksLdySg4Pn5KpGK7dKVhSfV3oxIdCkyyKS0"
MANGO_SHEET_GID = "1710203054"   # 시트 탭 gid (워크시트 이름으로 찾을 수 없을 때 폴백용)

# ── 망고카 시트 컬럼 매핑 ──────────────────────────────────────────────────────
# 시트 구조에 맞게 열 문자를 조정하세요
# 확인된 실제 열 구조
# A: 회원명  B: 회사명  C: 모델  D: 등급  E: 연료  F: 배기량
# G: 차대번호  H: 연식  I: 주행거리  J: A/M  K: 매물등록일
# L: 최종수정일  M: 구매요청일  N: 판매완료일  O: 업로드링크
# P: 매물상태  Q: 차량광고가
COL_SOURCE   = "A"   # 회원명 ("망고카지지오토 api" 필터)
COL_STATUS   = "P"   # 매물상태 ("게시" 필터)
COL_CODE     = "G"   # 차대번호 — 식별자로 사용
COL_CARNAME  = "C"   # 모델
COL_FUEL     = "E"   # 연료
COL_DISPLACE = "F"   # 배기량
COL_VIN      = "G"   # 차대번호
COL_YEAR     = "H"   # 연식
COL_MILEAGE  = "I"   # 주행거리
COL_TRANS    = "J"   # A/M (변속기)
COL_COLOR    = ""    # 없음
COL_LOCATION = ""    # 없음
COL_PRICE    = "Q"   # 차량광고가
COL_OPTIONS  = ""    # 없음
COL_SELLER   = "B"   # 회사명
COL_DRIVE    = "O"   # 업로드링크

SOURCE_FILTER = "망고카지지오토 api"   # A열 일치 조건
STATUS_FILTER  = "게시"               # P열 일치 조건
SHEET_START_ROW = 2                   # 헤더 다음 행부터 읽기


# ── 필드 변환 유틸 ────────────────────────────────────────────────────────────

def _strip_price(raw: str) -> str:
    """'$12,000' → '12000'"""
    if not raw:
        return ""
    cleaned = re.sub(r"[^\d.]", "", str(raw))
    if not cleaned:
        return ""
    try:
        return str(int(float(cleaned)))
    except ValueError:
        return ""


def _apply_beforward_fee(price_usd: int) -> int:
    """망고 USD 가격에 비포워드 업로드용 수수료 가산.

    구간(USD) → 가산액(USD):
        ≤1000:263, ≤1500:278, ≤2000:283, ≤3000:303,
        ≤5000:358, ≤6000:388, ≤7000:410, ≤8000:439,
        ≤10000:495, ≤15000:630, ≤20000:739, 그 외: 5%
    """
    if price_usd <= 1000:    fee = 263
    elif price_usd <= 1500:  fee = 278
    elif price_usd <= 2000:  fee = 283
    elif price_usd <= 3000:  fee = 303
    elif price_usd <= 5000:  fee = 358
    elif price_usd <= 6000:  fee = 388
    elif price_usd <= 7000:  fee = 410
    elif price_usd <= 8000:  fee = 439
    elif price_usd <= 10000: fee = 495
    elif price_usd <= 15000: fee = 630
    elif price_usd <= 20000: fee = 739
    else:                    fee = round(price_usd * 0.05)
    return price_usd + fee


def _normalize_mileage(raw: str) -> str:
    """'50,000KM' → '50000'"""
    if not raw:
        return ""
    digits = re.sub(r"[^\d]", "", str(raw))
    return digits if digits else ""


def _normalize_displacement(raw: str) -> str:
    """'2,400cc' → '2400'  숫자만 추출 (비포워드 필수 입력값)"""
    if not raw:
        return ""
    digits = re.sub(r"[^\d]", "", str(raw))
    return digits if digits else ""


def _normalize_fuel(raw: str) -> str:
    """GASOLINE / DIESEL / LPG / ELECTRIC / HYBRID → 비포워드 FUEL_MAP 키"""
    if not raw:
        return ""
    return raw.strip().lower()


def _normalize_transmission(raw: str) -> str:
    """AUTO / MANUAL / DCT / CVT → 비포워드 TRANSMISSION_MAP 키"""
    if not raw:
        return ""
    mapping = {
        "AUTO": "automatic",
        "AUTOMATIC": "automatic",
        "MANUAL": "manual",
        "DCT": "dct",
        "CVT": "cvt",
    }
    return mapping.get(raw.strip().upper(), raw.strip().lower())


_COLOR_EN_PATCH = {
    "navy":       "네이비",       # → 블루
    "pearl":      "펄",           # → 펄
    "champagne":  "샴페인골드",   # → 골드
    "grey":       "gray",         # grey → gray (둘 다 있지만 통일)
}

def _normalize_color(raw: str) -> str:
    """영문 색상명 → 비포워드 COLOR_MAP_KO 키로 변환."""
    if not raw:
        return ""
    lower = raw.strip().lower()
    return _COLOR_EN_PATCH.get(lower, lower)


def _extract_color_from_driver(driver) -> str:
    """이미 로드된 페이지에서 색상값 추출."""
    # 1) 확정 XPath
    try:
        el = driver.find_element(
            "xpath",
            "/html/body/main/div/div/section/div[1]/div[4]/div[4]/div[4]/span[2]/span"
        )
        val = el.text.strip()
        if val:
            return val
    except Exception:
        pass

    # 2) 색상명 키워드 폴백
    try:
        body_text = driver.find_element("tag name", "body").text
        m = re.search(
            r'\b(BLACK|WHITE|SILVER|GRAY|GREY|RED|BLUE|BROWN|GOLD|GREEN|'
            r'ORANGE|YELLOW|PURPLE|BEIGE|NAVY|PEARL|CHAMPAGNE)\b',
            body_text, re.IGNORECASE)
        if m:
            return m.group().upper()
    except Exception:
        pass

    return ""


def _fetch_colors_batch(detail_rows: list[dict]) -> None:
    """헤드리스 Chrome으로 망고 상세 페이지를 순회하며 색상 일괄 수집.
    결과는 각 row의 '색상' 키에 직접 저장.
    """
    import undetected_chromedriver as uc

    targets = [r for r in detail_rows if not r.get("색상") and r.get("_drive_link")]
    if not targets:
        return

    log.info("색상 크롤링 시작: %d건", len(targets))

    opts = uc.ChromeOptions()
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--window-size=1280,800")
    opts.add_argument("--window-position=-32000,-32000")  # 화면 밖으로 이동
    opts.add_argument("--lang=ko-KR")

    driver = uc.Chrome(options=opts, version_main=147)
    driver.set_page_load_timeout(30)

    try:
        for i, row in enumerate(targets, 1):
            url = row["_drive_link"].strip()
            code = url.split("/")[-1]
            try:
                driver.get(url)
                # JS 렌더링 대기 + 스크롤
                time.sleep(3)
                driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                time.sleep(1.5)
                driver.execute_script("window.scrollTo(0, 0);")

                color = _extract_color_from_driver(driver)
                if color:
                    row["색상"] = color
                    log.info("  [%d/%d] %s → 색상: %s", i, len(targets), code, color)
                else:
                    log.warning("  [%d/%d] %s → 색상 미수집", i, len(targets), code)

            except Exception as e:
                log.warning("  [%d/%d] 색상 크롤링 오류 (%s): %s", i, len(targets), code, e)
    finally:
        try:
            driver.quit()
        except Exception:
            pass


# ── 망고카 옵션 → 비포워드 옵션 매핑 ─────────────────────────────────────────

def _normalize_option_name(raw: str) -> str:
    """옵션명 비교용 정규화: 공백/구분문자 제거, 소문자화"""
    if raw is None:
        return ""
    return re.sub(r"[\s\-_·/()]+", "", str(raw).strip()).lower()


_MANGO_OPTION_MAP_CACHE: dict[str, str] | None = None


def _load_mango_option_map() -> dict[str, str]:
    """엑셀 '망고카' 시트 A열(망고카) → B열(비포워드) 매핑 로드"""
    global _MANGO_OPTION_MAP_CACHE
    if _MANGO_OPTION_MAP_CACHE is not None:
        return _MANGO_OPTION_MAP_CACHE

    mapping: dict[str, str] = {}
    try:
        from openpyxl import load_workbook
        wb = load_workbook(OPTION_MAPPING_FILE, data_only=True, read_only=True)
        ws = wb[OPTION_MAPPING_SHEET] if OPTION_MAPPING_SHEET in wb.sheetnames else wb.active
        for row in ws.iter_rows(min_row=2, min_col=1, max_col=2, values_only=True):
            mango_name = str(row[0]).strip() if row and row[0] is not None else ""
            beforward_name = str(row[1]).strip() if row and row[1] is not None else ""
            if mango_name and beforward_name:
                mapping[_normalize_option_name(mango_name)] = beforward_name
        wb.close()
        log.info("망고카 옵션 매핑 로드: %d개", len(mapping))
    except Exception as e:
        log.warning("망고카 옵션 매핑 로드 실패: %s", e)

    _MANGO_OPTION_MAP_CACHE = mapping
    return mapping


def _map_mango_options(raw_options: list[str]) -> list[OptionItem]:
    """망고카 추출 옵션 → 엑셀 매핑표 기준 비포워드 OptionItem 리스트"""
    option_map = _load_mango_option_map()
    mapped: list[OptionItem] = []
    seen: set[str] = set()
    unmapped: list[str] = []

    for raw in raw_options:
        mango_name = str(raw).strip()
        if not mango_name:
            continue
        bf_name = option_map.get(_normalize_option_name(mango_name))
        if not bf_name:
            unmapped.append(mango_name)
            continue
        if bf_name in seen:
            continue
        mapped.append(OptionItem(name=mango_name, mapped_name=bf_name))
        seen.add(bf_name)

    if mapped:
        log.info("  옵션 매핑: %d/%d 건 매핑됨", len(mapped), len(raw_options))
    if unmapped:
        log.info("  옵션 매핑 없음: %s", ", ".join(unmapped[:10]))
    return mapped


# 영문 모델명 → 재원표 한국어 키워드 매핑
_EN_TO_KO_MODEL = {
    "Grandeur":  "그랜저",
    "Sonata":    "소나타",
    "Avante":    "아반떼",
    "Elantra":   "아반떼",
    "Santa Fe":  "싼타페",
    "Tucson":    "투싼",
    "Ioniq":     "아이오닉",
    "Kona":      "코나",
    "Casper":    "캐스퍼",
    "Palisade":  "팰리세이드",
    "Sportage":  "스포티지",
    "Carnival":  "카니발",
    "Stinger":   "스팅어",
    "Seltos":    "셀토스",
    "Morning":   "모닝",
    "Soul":      "소울",
    "Spark":     "스파크",
    "SM6":       "SM6",
    "Ray":       "레이",
    "Ray ":      "레이",      # 뒤 공백으로 'Array' 오매칭 방지
    "Sorento":   "쏘렌토",
    "Mohave":    "모하비",
    "Starex":    "스타렉스",
    "Porter":    "포터",
    "Bongo":     "봉고",
    "Cruze":     "크루즈",
    "Malibu":    "말리부",
    "Trax":      "트랙스",
    "Captiva":   "캡티바",
    "Equinox":   "이쿼녹스",
}


def _append_korean_model(car_type: str) -> str:
    """영문 차종명에 재원표 한국어 키워드를 추가 (이미 있으면 그대로)."""
    ct_lower = car_type.lower()
    for en, ko in _EN_TO_KO_MODEL.items():
        if en.lower() in ct_lower and ko not in car_type:
            return f"{car_type} {ko}"
    return car_type


def _make_car_info(detail: dict) -> CarInfo:
    """
    mango_crawler 상세 dict → CarInfo 객체 변환.

    mango 데이터 키:
        상품코드, 차량명, 연식, 가격(USD), 변속기, 색상, 배기량,
        구동방식, 연료타입, 최초등록일, 위치, 차대번호, 차대번호_상세,
        주행거리(상세), 보유옵션(전체), 판매자
    """
    info = CarInfo()

    # 차종명 — 비포워드 재원표 조회에 사용 (가장 중요)
    # 영문 공개 페이지에서 수집된 경우 한국어 키워드를 추가해 재원표 매칭 보정
    info.car_type = _append_korean_model((detail.get("차량명") or "").strip())

    # 연식 (예: "2020")
    yr = detail.get("연식", "")
    if yr:
        info.year_month = yr.replace("년", "").strip()

    # 주행거리
    info.mileage = _normalize_mileage(detail.get("주행거리(상세)", ""))

    # 배기량
    info.displacement = _normalize_displacement(detail.get("배기량", ""))

    # 연료
    info.fuel_type = _normalize_fuel(detail.get("연료타입", ""))

    # 변속기
    info.transmission = _normalize_transmission(detail.get("변속기", ""))

    # 색상 — _fetch_colors_batch 에서 미리 채워진 값 사용
    info.color = _normalize_color(detail.get("색상", ""))

    # 위치
    info.location = (detail.get("위치") or "").strip()

    # 가격 (USD 숫자만) — 비포워드 업로드용 수수료 가산
    raw_price = _strip_price(detail.get("가격(USD)", ""))
    if raw_price:
        info.price = str(_apply_beforward_fee(int(raw_price)))
        log.info("  가격 변환: $%s → $%s (수수료 가산)", raw_price, info.price)
    else:
        info.price = ""

    # 차대번호 — 상세 페이지 우선, 없으면 목록 데이터
    vin = (detail.get("차대번호_상세") or detail.get("차대번호") or "").strip()
    info.inspection_chassis_no = vin

    # 옵션 리스트 — 망고카 옵션을 엑셀 매핑표에 따라 비포워드 옵션으로 변환
    opts_raw = detail.get("보유옵션(전체)", "")
    if opts_raw:
        raw_list = [o.strip() for o in opts_raw.split("/") if o.strip()]
        info.options = _map_mango_options(raw_list)
    else:
        info.options = []

    # 이미지 소스 — O열(업로드링크)의 URL을 그대로 사용
    drive_link = (detail.get("_drive_link") or "").strip()
    if drive_link:
        info.drive_link = drive_link

    return info


# ── 수집 단계 ─────────────────────────────────────────────────────────────────

def _col_to_idx(col: str) -> int:
    """열 문자 → 0-based 인덱스 (A=0, Z=25, AA=26 …)"""
    idx = 0
    for c in col.upper():
        idx = idx * 26 + (ord(c) - ord("A") + 1)
    return idx - 1


def _get_service_account_file() -> str:
    """현재 디렉토리에서 서비스 계정 JSON 파일 탐색"""
    candidate = HERE / "adjustmentdata-51a7199ac3ba.json"
    if candidate.exists():
        return str(candidate)
    raise FileNotFoundError(
        "서비스 계정 JSON 파일을 찾을 수 없습니다.\n"
        f"'{candidate}' 위치에 파일을 복사해 주세요."
    )


def collect_from_sheets() -> list[dict]:
    """
    Google Sheets 망고카 통합 시트에서 지지오토 게시 차량 수집.

    필터 조건:
        A열 == SOURCE_FILTER ("망고카지지오토 api")
        P열 == STATUS_FILTER  ("게시")

    반환값: _make_car_info() 가 인식하는 dict 리스트
    """
    log.info("Google Sheets 데이터 수집 시작 (ID=%s)", MANGO_SHEET_ID)

    try:
        import gspread
        from google.oauth2.service_account import Credentials
    except ImportError:
        log.error("gspread / google-auth 패키지가 없습니다. pip install gspread google-auth")
        return []

    sa_file = _get_service_account_file()
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets.readonly",
        "https://www.googleapis.com/auth/drive.readonly",
    ]
    creds = Credentials.from_service_account_file(sa_file, scopes=scopes)
    gc = gspread.authorize(creds)

    spreadsheet = gc.open_by_key(MANGO_SHEET_ID)

    # gid로 워크시트 찾기
    worksheet = None
    for ws in spreadsheet.worksheets():
        if str(ws.id) == MANGO_SHEET_GID:
            worksheet = ws
            break
    if worksheet is None:
        log.warning("gid=%s 시트 없음 → 첫 번째 시트 사용", MANGO_SHEET_GID)
        worksheet = spreadsheet.sheet1

    log.info("시트 '%s' 전체 읽기 중...", worksheet.title)
    all_values = worksheet.get_all_values()
    if len(all_values) < SHEET_START_ROW:
        log.warning("시트에 데이터 행이 없습니다.")
        return []

    detail_rows: list[dict] = []
    skipped_source = 0
    skipped_status = 0

    for row_num, row in enumerate(all_values[SHEET_START_ROW - 1:], start=SHEET_START_ROW):
        def cell(col: str) -> str:
            i = _col_to_idx(col)
            return row[i].strip() if i < len(row) else ""

        source_val = cell(COL_SOURCE)
        status_val = cell(COL_STATUS)

        if source_val != SOURCE_FILTER:
            skipped_source += 1
            continue
        if status_val != STATUS_FILTER:
            skipped_status += 1
            continue

        d: dict = {
            "상품코드":       cell(COL_CODE),
            "차량명":         cell(COL_CARNAME),
            "연식":           cell(COL_YEAR),
            "주행거리(상세)": cell(COL_MILEAGE),
            "배기량":         cell(COL_DISPLACE),
            "연료타입":       cell(COL_FUEL),
            "변속기":         cell(COL_TRANS),
            "색상":           cell(COL_COLOR),
            "위치":           cell(COL_LOCATION),
            "가격(USD)":      cell(COL_PRICE),
            "차대번호":       cell(COL_VIN),
            "차대번호_상세":  cell(COL_VIN),
            "보유옵션(전체)": cell(COL_OPTIONS),
            "판매자":         cell(COL_SELLER),
            "_drive_link":    cell(COL_DRIVE),  # 이미지 드라이브 링크 (내부용)
            "_sheet_row":     row_num,
        }
        detail_rows.append(d)

    log.info(
        "시트 필터 결과: %d건 수집 (구분불일치=%d, 상태불일치=%d)",
        len(detail_rows), skipped_source, skipped_status,
    )
    return detail_rows


# ── 업로드 단계 ───────────────────────────────────────────────────────────────

def upload_to_beforward(detail_rows: list[dict]) -> tuple[int, int]:
    """
    CarInfo 로 변환 후 비포워드 업로드.
    Returns:
        (총시도, 성공)
    """
    if not detail_rows:
        log.warning("업로드할 데이터 없음")
        return 0, 0

    # 색상 없는 차량 → 망고 페이지에서 일괄 크롤링
    _fetch_colors_batch(detail_rows)

    uploader = BefowordCrawler(headless=False)

    log.info("비포워드 로그인...")
    if not uploader.login():
        log.error("비포워드 로그인 실패")
        return 0, 0

    total = 0
    success = 0

    for i, detail in enumerate(detail_rows, 1):
        code = detail.get("상품코드", f"ROW_{i}")
        log.info("[%d/%d] 업로드: %s", i, len(detail_rows), code)

        car_info = _make_car_info(detail)

        if not car_info.car_type:
            log.warning("  [SKIP] 차량명 없음 → 건너뜀")
            continue

        if not car_info.inspection_chassis_no:
            log.warning("  [SKIP] 차대번호 없음 → 건너뜀 (%s)", code)
            continue

        if not car_info.price:
            log.warning("  [SKIP] 가격 없음 → 건너뜀 (%s)", code)
            continue

        # ── 안전장치: 지지오토 매물만 업로드 (필터 누수 방지) ──
        seller = (detail.get("판매자") or "").strip()
        if not re.search(r"GG[-\s]?AUTO|지지오토", seller, re.IGNORECASE):
            log.warning("  [SKIP] 판매자 미확인 → 업로드 차단 (%s, 판매자='%s')",
                        code, seller)
            continue

        log.info("  차종: %s | 연식: %s | 주행: %s | 가격: $%s | VIN: %s",
                 car_info.car_type, car_info.year_month, car_info.mileage,
                 car_info.price, car_info.inspection_chassis_no)

        total += 1
        try:
            ok = uploader.fill_vehicle_data(car_info, auto_submit=not DRY_RUN)

            submitted = getattr(uploader, '_listing_submitted', False)
            if DRY_RUN and ok and not submitted:
                log.info("  [DRY_RUN] form filled only; not submitted")
            elif ok or submitted:
                listing_id = getattr(car_info, '_listing_id', '')
                log.info("  [OK] 업로드 성공 | ID=%s", listing_id)
                success += 1
            else:
                step  = getattr(uploader, '_last_error_step', '미캡처')
                cause = getattr(uploader, '_last_error_cause', '')
                log.error("  [FAIL] 단계: %s | 원인: %s", step, cause)

        except Exception as e:
            log.error("  [ERROR] %s\n%s", e, traceback.format_exc())

            # Chrome 세션 복구
            err_s = str(e).lower()
            if any(k in err_s for k in ('invalid session', 'session', 'crashed', 'disconnected')):
                log.info("  Chrome 세션 복구 시도...")
                try:
                    uploader.close()
                except Exception:
                    pass
                time.sleep(2)
                uploader._setup_driver()
                if not uploader.login():
                    log.error("  재로그인 실패 — 업로드 중단")
                    break

        # 이미지 임시파일 정리
        downloads = getattr(uploader, '_last_downloaded_image_files', [])
        if downloads:
            try:
                uploader._cleanup_downloaded_images(downloads)
            except Exception:
                pass
            uploader._last_downloaded_image_files = []

        time.sleep(2)

    try:
        uploader.close()
    except Exception:
        pass

    return total, success


# ── 판매 완료 감지 + 게시 정지 ───────────────────────────────────────────────

# 망고 공개 상세 페이지에서 판매 완료 여부를 판단하는 XPath
SOLD_XPATH = "/html/body/div[1]/div/div/div[2]/p"


def _is_sold_on_mango(driver, code: str) -> bool:
    """망고 공개 상세 페이지에서 판매 완료 여부 확인.

    XPath SOLD_XPATH 가 존재하면 판매 완료로 판단.
    """
    url = f"{DETAIL_BASE}/{code}"
    try:
        driver.get(url)
        time.sleep(1.5)
        elems = driver.find_elements("xpath", SOLD_XPATH)
        return len(elems) > 0
    except Exception as e:
        log.warning("  [WARN] 판매여부 확인 실패 (%s): %s", code, e)
        return False


def suspend_sold_vehicles(detail_rows: list[dict]) -> tuple[list[str], int]:
    """수집된 차량 중 판매 완료된 것을 비포워드에서 게시 정지.

    Args:
        detail_rows: collect_mango_data() 반환값

    Returns:
        (판매완료_VIN_목록, 게시정지_성공수)
    """
    if not detail_rows:
        log.warning("점검할 차량 데이터 없음")
        return [], 0

    import importlib.util
    spec = importlib.util.spec_from_file_location(
        "mango_crawler", HERE / "mango_crawler.py"
    )
    mc = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mc)

    driver = mc.make_driver()
    sold_codes: list[tuple[str, str]] = []  # (code, vin)

    try:
        for row in detail_rows:
            code = row.get("상품코드", "")
            vin = (row.get("차대번호_상세") or row.get("차대번호") or "").strip()
            if not code or not vin:
                continue
            if _is_sold_on_mango(driver, code):
                log.info("  [SOLD] %s (VIN: %s)", code, vin)
                sold_codes.append((code, vin))
            else:
                log.info("  [ON SALE] %s", code)
    finally:
        driver.quit()

    sold_vins = [vin for _, vin in sold_codes]

    if not sold_codes:
        log.info("판매 완료 차량 없음 — 게시 정지 불필요")
        return [], 0

    log.info("판매 완료 %d건 → 비포워드 게시 정지 시작", len(sold_codes))

    suspender = BeforwardSuspensionManager(headless=False)
    if not suspender.login():
        log.error("비포워드 로그인 실패 — 게시 정지 중단")
        return sold_vins, 0

    suspended = 0
    for code, vin in sold_codes:
        log.info("  게시 정지 시도: %s (VIN: %s)", code, vin)
        try:
            ok = suspender.search_and_delete_listing(vin)
            if ok:
                log.info("  [OK] 게시 정지 완료: %s", code)
                suspended += 1
            else:
                log.warning("  [FAIL] 게시 정지 실패: %s", code)
        except Exception as e:
            log.error("  [ERROR] %s: %s", code, e)
        time.sleep(1)

    try:
        suspender.close()
    except Exception:
        pass

    return sold_vins, suspended


# ── 메인 ─────────────────────────────────────────────────────────────────────

def main():
    log.info("=" * 60)
    log.info("망고월드카 → 비포워드 자동 업로드 시작")
    log.info("=" * 60)

    # 1. 수집 (Google Sheets → 망고카지지오토 api + 게시 필터)
    detail_rows = collect_from_sheets()

    if not detail_rows:
        log.warning("수집된 차량 없음. 종료.")
        return

    # 2. 판매 완료 차량 → 게시 정지 (TODO: 시트 기반 감지로 재구현 예정)

    # 3. 업로드
    total, success = upload_to_beforward(detail_rows)

    log.info("=" * 60)
    log.info("완료: 업로드 %d/%d 성공", success, total)
    log.info("=" * 60)


if __name__ == "__main__":
    main()
