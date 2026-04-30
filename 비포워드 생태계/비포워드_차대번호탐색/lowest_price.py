import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import time
import os
import random
import re
import threading

# === 설정 ===
TARGET_PATH = os.path.join(os.path.dirname(__file__), '게시중 엔카 차량_비포워드 금액비교_결과.xlsx')
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), '게시중 엔카 차량_비포워드 금액비교_최저가.xlsx')
TEMP_PATH = os.path.join(os.path.dirname(__file__), '게시중 엔카 차량_비포워드 금액비교_최저가.tmp')
SEARCH_URL = 'https://www.beforward.jp/stocklist/sortkey=n/keyword={vin}/kmode=and/'
SAVE_INTERVAL = 30
NUM_WORKERS = 2


def safe_save(wb):
    wb.save(TEMP_PATH)
    os.replace(TEMP_PATH, OUTPUT_PATH)


def setup_driver():
    options = Options()
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-blink-features=AutomationControlled')
    options.add_experimental_option('excludeSwitches', ['enable-automation'])
    options.add_experimental_option('useAutomationExtension', False)
    options.add_argument('user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
    options.page_load_strategy = 'eager'
    driver = webdriver.Chrome(options=options)
    driver.execute_cdp_cmd('Page.addScriptToEvaluateOnNewDocument', {
        'source': 'Object.defineProperty(navigator, "webdriver", {get: () => undefined})'
    })
    driver.set_page_load_timeout(15)
    return driver


def is_blocked(driver):
    try:
        src = driver.page_source.lower()
        if len(src) < 3000:
            return True
        if 'blocked' in src:
            return True
        if 'err_' in src or 'net::' in src:
            return True
        if 'beforward' not in src:
            return True
        return False
    except Exception:
        return True


def parse_price(txt):
    """'$1,234' 형태에서 숫자 추출"""
    try:
        return float(re.sub(r'[^\d.]', '', txt.replace(',', '')))
    except Exception:
        return None


def search_vin_lowest(driver, vin):
    driver.get(SEARCH_URL.format(vin=vin))
    time.sleep(random.uniform(2.0, 3.5))

    if is_blocked(driver):
        raise Exception('IP_BLOCKED')

    try:
        meta = driver.find_element(By.XPATH, '//meta[@name="ga_stocklist_results"]')
        result_count = int(meta.get_attribute('content'))
    except Exception:
        result_count = 0

    if result_count == 0:
        return '없음'

    # 모든 span.price 수집 → 최저가 반환
    try:
        price_elements = driver.find_elements(By.CSS_SELECTOR, 'span.price')
        candidates = []
        for p in price_elements:
            txt = p.text.strip()
            if txt and '$' in txt:
                num = parse_price(txt)
                if num and num > 0:
                    candidates.append((num, txt))
        if candidates:
            if len(candidates) > 1:
                print(f'  [{vin}] 여러건이 나왔다! ({len(candidates)}개) → 최저가 선택')
            return min(candidates, key=lambda x: x[0])[1]
    except Exception:
        pass

    return '없음'


def worker(worker_id, task_queue, ws, wb, lock, block_event, counter, total, start_time):
    driver = setup_driver()
    for _ in range(3):
        try:
            driver.get('https://www.beforward.jp/')
            break
        except Exception as e:
            print(f'[워커{worker_id}] 초기 접속 실패, 재시도: {e}')
            try:
                driver.quit()
            except Exception:
                pass
            time.sleep(5)
            driver = setup_driver()
    time.sleep(2 + worker_id)
    print(f'[워커{worker_id}] 시작: {driver.title}')

    try:
        while True:
            if not block_event.is_set():
                print(f'[워커{worker_id}] 차단 대기 중...')
                block_event.wait()
                try:
                    driver.get('https://www.beforward.jp/')
                    time.sleep(3)
                except Exception:
                    try:
                        driver.quit()
                    except Exception:
                        pass
                    try:
                        driver = setup_driver()
                        driver.get('https://www.beforward.jp/')
                        time.sleep(3)
                    except Exception:
                        pass

            with lock:
                if not task_queue:
                    break
                row, vin, old_price = task_queue.pop(0)

            try:
                price = search_vin_lowest(driver, vin)
            except Exception as e:
                if 'IP_BLOCKED' in str(e):
                    if block_event.is_set():
                        print(f'\n[워커{worker_id}] ★ IP 차단 감지 - 전체 워커 일시정지. 5분 후 자동 재개됩니다 ★')
                        block_event.clear()
                        with lock:
                            safe_save(wb)
                        time.sleep(300)
                        print(f'[워커{worker_id}] 5분 경과 - 재접속 시도...')
                        block_event.set()
                    else:
                        block_event.wait()

                    with lock:
                        task_queue.insert(0, (row, vin, old_price))
                    continue
                else:
                    print(f'[워커{worker_id}] 오류 {vin}: {e}')
                    price = '오류'
                    if any(k in str(e).lower() for k in ('renderer', 'session', 'target window', 'err_', 'network', 'access_denied', 'disconnected')):
                        print(f'[워커{worker_id}] 드라이버 재시작...')
                        try:
                            driver.quit()
                        except Exception:
                            pass
                        try:
                            driver = setup_driver()
                            driver.get('https://www.beforward.jp/')
                            time.sleep(3)
                        except Exception as re_e:
                            print(f'[워커{worker_id}] 재시작 실패: {re_e}')
                            time.sleep(5)
                        with lock:
                            task_queue.insert(0, (row, vin, old_price))
                        continue

            with lock:
                ws.cell(row, 10).value = price
                counter[0] += 1
                done = counter[0]
                elapsed = time.time() - start_time
                per_item = elapsed / done
                remaining = (total - done) * per_item / 60

                # 기존 가격과 비교 후 콘솔 출력
                old_num = parse_price(old_price)
                new_num = parse_price(price) if '$' in str(price) else None
                if old_num and new_num and new_num < old_num:
                    diff = old_num - new_num
                    tag = f'↓ 더 낮음! 기존 {old_price} → {price} (-${diff:,.0f})'
                else:
                    tag = price
                print(f'[워커{worker_id}][{done}/{total}] Row{row}: {vin} -> {tag}  ({per_item:.1f}s/개, 남은:{remaining:.0f}분)')

                if done % SAVE_INTERVAL == 0:
                    safe_save(wb)
                    print(f'  === {done}개 완료, 중간 저장 ===')
    except Exception as e:
        print(f'[워커{worker_id}] 예외 종료: {e}')
    finally:
        driver.quit()
        print(f'[워커{worker_id}] 종료')


def main():
    if not os.path.exists(TARGET_PATH):
        print(f'결과 파일 없음: {TARGET_PATH}')
        return

    print(f'파일 로드: {TARGET_PATH}')
    wb = openpyxl.load_workbook(TARGET_PATH)
    ws = wb['남미주요차종']
    max_row = ws.max_row

    # J열에 $ 가격이 있는 행만 재크롤링 대상 (기존 가격도 함께 저장)
    task_queue = []
    for row in range(2, max_row + 1):
        j_val = str(ws.cell(row, 10).value or '').strip()
        if '$' not in j_val:
            continue
        vin = ws.cell(row, 3).value
        if not vin or str(vin).strip() == '':
            continue
        task_queue.append((row, str(vin).strip(), j_val))  # (행, VIN, 기존가격)

    total = len(task_queue)
    print(f'처리 대상 (가격 있는 행): {total}개')

    if total == 0:
        print('처리할 항목 없음')
        return

    lock = threading.Lock()
    block_event = threading.Event()
    block_event.set()
    counter = [0]
    start_time = time.time()

    threads = []
    for i in range(NUM_WORKERS):
        t = threading.Thread(
            target=worker,
            args=(i + 1, task_queue, ws, wb, lock, block_event, counter, total, start_time),
            daemon=True
        )
        threads.append(t)

    try:
        for t in threads:
            t.start()
        for t in threads:
            t.join()
    except KeyboardInterrupt:
        print('\n사용자에 의해 중단됨')
    finally:
        safe_save(wb)
        total_time = time.time() - start_time
        print(f'\n최종 저장 완료! 총 {counter[0]}개 처리, 소요시간: {total_time / 60:.1f}분')


if __name__ == '__main__':
    while True:
        try:
            main()
            break
        except KeyboardInterrupt:
            print('\n사용자에 의해 중단됨')
            break
        except Exception as e:
            print(f'\n[재시작] 오류로 중단됨: {e}')
            print('10초 후 재시작...')
            time.sleep(10)
