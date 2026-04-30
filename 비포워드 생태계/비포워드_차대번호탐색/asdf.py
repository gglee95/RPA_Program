import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import time
import os
import random
import threading

# === 설정 ===
EXCEL_PATH = os.path.join(os.path.dirname(__file__), '게시중 엔카 차량_비포워드 금액비교.xlsx')
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), '게시중 엔카 차량_비포워드 금액비교_결과.xlsx')
TEMP_PATH = os.path.join(os.path.dirname(__file__), '게시중 엔카 차량_비포워드 금액비교_결과.tmp')
SEARCH_URL = 'https://www.beforward.jp/stocklist/sortkey=n/keyword={vin}/kmode=and/'
SAVE_INTERVAL = 30
NUM_WORKERS =3


def safe_save(wb):
    wb.save(TEMP_PATH)
    os.replace(TEMP_PATH, OUTPUT_PATH)


def setup_driver():
    options = Options()
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-blink-features=AutomationControlled')
    options.add_experimental_option('excludeSwitches', ['enable-automation'])
    options.add_experimental_option('useAutomationExtension', False)
    options.add_argument('user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
    options.page_load_strategy = 'eager'
    driver = webdriver.Chrome(options=options)
    driver.execute_cdp_cmd('Page.addScriptToEvaluateOnNewDocument', {
        'source': 'Object.defineProperty(navigator, "webdriver", {get: () => undefined})'
    })
    driver.set_page_load_timeout(15)
    return driver


def is_blocked(driver):
    try:
        src = driver.page_source.lower()
        if len(src) < 3000:
            return True
        if 'blocked' in src:
            return True
        if 'err_' in src or 'net::' in src:
            return True
        if 'beforward' not in src:
            return True
        return False
    except Exception:
        return True


def search_vin(driver, vin):
    url = SEARCH_URL.format(vin=vin)
    driver.get(url)
    time.sleep(random.uniform(2.0, 3.5))

    if is_blocked(driver):
        raise Exception('IP_BLOCKED')

    try:
        meta = driver.find_element(By.XPATH, '//meta[@name="ga_stocklist_results"]')
        result_count = int(meta.get_attribute('content'))
    except Exception:
        result_count = 0

    if result_count == 0:
        return '없음'

    try:
        target = driver.find_element(By.XPATH,
            '//*[@id="list-content"]/div[3]/div[1]/table/tbody/tr/td[3]/div/a/div/div[1]/p[2]/span[2]')
        if target.text.strip():
            return target.text.strip()
    except Exception:
        pass

    try:
        prices = driver.find_elements(By.CSS_SELECTOR, 'span.price')
        for p in prices:
            txt = p.text.strip()
            if txt and '$' in txt:
                return txt
    except Exception:
        pass

    return '없음'


def worker(worker_id, task_queue, ws, wb, lock, block_event, counter, total, start_time):
    driver = setup_driver()
    for _ in range(3):
        try:
            driver.get('https://www.beforward.jp/')
            break
        except Exception as e:
            print(f'[워커{worker_id}] 초기 접속 실패, 재시도: {e}')
            try:
                driver.quit()
            except Exception:
                pass
            time.sleep(5)
            driver = setup_driver()
    time.sleep(2 + worker_id)
    print(f'[워커{worker_id}] 시작: {driver.title}')

    try:
        while True:
            # 전체 차단 중이면 해제될 때까지 대기 (set=정상, clear=차단중)
            if not block_event.is_set():
                print(f'[워커{worker_id}] 차단 대기 중...')
                block_event.wait()  # set될 때까지(=정상 복귀까지) 대기
                try:
                    driver.get('https://www.beforward.jp/')
                    time.sleep(3)
                except Exception:
                    try:
                        driver.quit()
                    except Exception:
                        pass
                    try:
                        driver = setup_driver()
                        driver.get('https://www.beforward.jp/')
                        time.sleep(3)
                    except Exception:
                        pass

            with lock:
                if not task_queue:
                    break
                row, vin = task_queue.pop(0)

            try:
                price = search_vin(driver, vin)
            except Exception as e:
                if 'IP_BLOCKED' in str(e):
                    if block_event.is_set():  # 아직 정상 상태 → 첫 차단 감지
                        print(f'\n[워커{worker_id}] ★ IP 차단 감지 - 전체 워커 일시정지. 5분 후 자동 재개됩니다 ★')
                        block_event.clear()  # 차단 표시 (다른 워커들 대기)
                        with lock:
                            safe_save(wb)
                        time.sleep(300)
                        print(f'[워커{worker_id}] 5분 경과 - 재접속 시도...')
                        block_event.set()  # 정상 복귀 (다른 워커들 재개)
                    else:
                        block_event.wait()  # 다른 워커가 해제할 때까지 대기

                    with lock:
                        task_queue.insert(0, (row, vin))
                    continue
                else:
                    print(f'[워커{worker_id}] 오류 {vin}: {e}')
                    price = '오류'
                    # 렌더러 오류면 드라이버 재시작
                    if any(k in str(e).lower() for k in ('renderer', 'session', 'target window', 'err_', 'network', 'access_denied', 'disconnected')):
                        print(f'[워커{worker_id}] 드라이버 재시작...')
                        try:
                            driver.quit()
                        except Exception:
                            pass
                        try:
                            driver = setup_driver()
                            driver.get('https://www.beforward.jp/')
                            time.sleep(3)
                        except Exception as re:
                            print(f'[워커{worker_id}] 재시작 실패, 재시도: {re}')
                            time.sleep(5)
                        with lock:
                            task_queue.insert(0, (row, vin))
                        continue

            with lock:
                ws.cell(row, 10).value = price
                counter[0] += 1
                done = counter[0]
                elapsed = time.time() - start_time
                per_item = elapsed / done
                remaining = (total - done) * per_item / 60
                print(f'[워커{worker_id}][{done}/{total}] Row{row}: {vin} -> {price}  ({per_item:.1f}s/개, 남은:{remaining:.0f}분)')

                if done % SAVE_INTERVAL == 0:
                    safe_save(wb)
                    print(f'  === {done}개 완료, 중간 저장 ===')
    except Exception as e:
        print(f'[워커{worker_id}] 예외 종료: {e}')
    finally:
        driver.quit()
        print(f'[워커{worker_id}] 종료')


def main():
    if os.path.exists(OUTPUT_PATH):
        print('이전 결과 파일 발견 - 이어서 진행')
        wb = openpyxl.load_workbook(OUTPUT_PATH)
    else:
        print('원본 파일 로드')
        wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb['남미주요차종']
    max_row = ws.max_row
    print(f'총 {max_row - 1}개 차량')

    task_queue = []
    for row in range(2, max_row + 1):
        j_val = ws.cell(row, 10).value
        if j_val is not None and str(j_val).strip() not in ('', '오류'):
            continue
        vin = ws.cell(row, 3).value
        if not vin or str(vin).strip() == '':
            ws.cell(row, 10).value = '없음'
            continue
        task_queue.append((row, str(vin).strip()))

    total = len(task_queue)
    print(f'처리 대상: {total}개')

    if total == 0:
        print('처리할 항목 없음')
        return

    lock = threading.Lock()
    block_event = threading.Event()  # set=정상, clear=차단중
    block_event.set()  # 초기값: 정상
    counter = [0]
    start_time = time.time()

    threads = []
    for i in range(NUM_WORKERS):
        t = threading.Thread(
            target=worker,
            args=(i + 1, task_queue, ws, wb, lock, block_event, counter, total, start_time),
            daemon=True
        )
        threads.append(t)

    try:
        for t in threads:
            t.start()
        for t in threads:
            t.join()
    except KeyboardInterrupt:
        print('\n사용자에 의해 중단됨')
    finally:
        safe_save(wb)
        total_time = time.time() - start_time
        print(f'\n최종 저장 완료! 총 {counter[0]}개 처리, 소요시간: {total_time / 60:.1f}분')


if __name__ == '__main__':
    while True:
        try:
            main()
            break
        except KeyboardInterrupt:
            print('\n사용자에 의해 중단됨')
            break
        except Exception as e:
            print(f'\n[재시작] 오류로 중단됨: {e}')
            print('10초 후 재시작...')
            time.sleep(10)
