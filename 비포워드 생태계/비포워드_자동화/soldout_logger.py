"""
Logging infrastructure for Encar SOLD OUT Monitoring System
Provides multi-channel logging (console + files) with structured messages
"""
import logging
from datetime import datetime
from pathlib import Path
from config import LOG_DIR

EXCEL_LOG_DIR = Path(r"C:\SynologyDrive\데이터 관련\비포워드 데이터 관련\비포워드 에러 로그")

# 진단 로그 포맷:
# 2026-04-10 11:30:00 | ERROR | Row 1460 | 모델: 뉴 A6 35 TDI | 단계: 제조사_선택 | 원인: 'AUDI' 옵션 없음


class SoldOutLogger:
    """Multi-channel logging for monitoring system"""

    def __init__(self):
        """Initialize loggers for success, errors, and history"""
        # Ensure log directory exists
        LOG_DIR.mkdir(exist_ok=True)

        # Success logger
        self.success_logger = self._create_logger(
            'success',
            LOG_DIR / 'soldout_success.log'
        )

        # Error logger
        self.error_logger = self._create_logger(
            'error',
            LOG_DIR / 'soldout_errors.log',
            level=logging.ERROR
        )

        # Suspension history logger
        self.history_logger = self._create_logger(
            'history',
            LOG_DIR / 'suspension_history.log'
        )

        # 진단 로거: 날짜별 파일, 모델/단계/원인 구조화
        today = datetime.now().strftime('%Y%m%d')
        self.diagnostic_logger = self._create_logger(
            f'diagnostic_{today}',
            LOG_DIR / f'errors_{today}.log',
            level=logging.DEBUG
        )

    def _create_logger(self, name: str, filepath: Path, level=logging.INFO):
        """Create logger with file and console handlers

        Args:
            name: Logger name
            filepath: Log file path
            level: Logging level

        Returns:
            Configured logger instance
        """
        logger = logging.getLogger(name)
        logger.setLevel(level)

        # Clear existing handlers to avoid duplicates
        logger.handlers.clear()

        # File handler
        fh = logging.FileHandler(filepath, encoding='utf-8')
        fh.setLevel(level)

        # Console handler
        ch = logging.StreamHandler()
        ch.setLevel(level)

        # Formatter with timestamp
        formatter = logging.Formatter(
            '%(asctime)s | %(levelname)s | %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S'
        )
        fh.setFormatter(formatter)
        ch.setFormatter(formatter)

        logger.addHandler(fh)
        logger.addHandler(ch)

        return logger

    def log_cycle_start(self, cycle_number: int):
        """Log monitoring cycle start

        Args:
            cycle_number: Current cycle number
        """
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        msg = f"Cycle {cycle_number} started at {timestamp}"
        self.success_logger.info(msg)
        self.history_logger.info(msg)

    def log_cycle_end(self, cycle_number: int, processed: int, suspended: int):
        """Log monitoring cycle completion

        Args:
            cycle_number: Current cycle number
            processed: Number of rows processed
            suspended: Number of suspensions performed
        """
        msg = f"Cycle {cycle_number} completed - Processed {processed} rows, {suspended} suspensions"
        self.success_logger.info(msg)
        self.history_logger.info(msg)

    def log_soldout_detected(self, row: int, url: str, method: str):
        """Log successful SOLD OUT detection

        Args:
            row: Row number in Google Sheets
            url: Encar URL
            method: Detection method used (text, redirect, etc.)
        """
        msg = f"Row {row} | SOLD OUT detected via {method} | URL: {url}"
        self.success_logger.info(msg)
        self.history_logger.info(msg)

    def log_suspension_success(self, row: int, car_number: str, listing_url: str):
        """Log successful listing suspension

        Args:
            row: Row number in Google Sheets
            car_number: Vehicle number or VIN
            listing_url: ByForward listing URL
        """
        msg = f"Row {row} | Vehicle {car_number} | Suspended: {listing_url}"
        self.success_logger.info(msg)
        self.history_logger.info(msg)

    def log_suspension_failed(self, row: int, car_number: str, reason: str):
        """Log failed suspension attempt

        Args:
            row: Row number in Google Sheets
            car_number: Vehicle number or VIN
            reason: Failure reason
        """
        msg = f"Row {row} | Vehicle {car_number} | Suspension failed: {reason}"
        self.error_logger.error(msg)
        self.history_logger.error(msg)

    def log_error(self, row: int, error_type: str, message: str):
        """Log general error

        Args:
            row: Row number (0 for system-level errors)
            error_type: Error category
            message: Error message
        """
        if row > 0:
            msg = f"Row {row} | {error_type} | {message}"
        else:
            msg = f"{error_type} | {message}"
        self.error_logger.error(msg)

    def log_info(self, message: str):
        """Log informational message

        Args:
            message: Info message
        """
        self.success_logger.info(message)

    def log_warning(self, row: int, message: str):
        """Log warning message

        Args:
            row: Row number
            message: Warning message
        """
        msg = f"Row {row} | WARNING | {message}"
        self.error_logger.warning(msg)

    def log_system_start(self):
        """Log system startup"""
        msg = "=" * 70 + "\n" + \
              "Encar SOLD OUT Monitoring System Started".center(70) + "\n" + \
              "=" * 70
        self.success_logger.info(msg)
        self.history_logger.info(msg)

    def log_system_shutdown(self):
        """Log system shutdown"""
        msg = "System shutdown initiated - Goodbye!"
        self.success_logger.info(msg)
        self.history_logger.info(msg)

    # ── 업로드 관련 ───────────────────────────────────────────────────────────

    def log_upload_start(self, row: int, encar_url: str):
        msg = f"Row {row} | 업로드 시작 | URL: {encar_url}"
        self.success_logger.info(msg)
        self.history_logger.info(msg)

    def log_upload_success(self, row: int, listing_id: str):
        msg = f"Row {row} | 업로드 성공 | 비포워드 ID: {listing_id}"
        self.success_logger.info(msg)
        self.history_logger.info(msg)

    def log_upload_failed(self, row: int, reason: str):
        msg = f"Row {row} | 업로드 실패 | {reason}"
        self.error_logger.error(msg)
        self.history_logger.error(msg)

    def log_image_download(self, row: int, source: str, count: int):
        """source: 'gdrive' or 'mango'"""
        msg = f"Row {row} | 이미지 다운로드 완료 | 소스: {source} | {count}개"
        self.success_logger.info(msg)

    def log_image_upload_success(self, row: int, listing_id: str, count: int):
        msg = f"Row {row} | 이미지 업로드 완료 | 비포워드 ID: {listing_id} | {count}개"
        self.success_logger.info(msg)
        self.history_logger.info(msg)

    def log_condition_form(self, row: int, is_4wd: bool):
        drive_str = "4WD" if is_4wd else "2WD"
        msg = f"Row {row} | condition-form 완료 | 구동방식: {drive_str}"
        self.success_logger.info(msg)

    # ── 매입완료 게시정지 관련 ────────────────────────────────────────────────

    def log_purchased_suspension(self, row: int, chassis_no: str):
        msg = f"Row {row} | 매입완료 게시정지 | 차대번호: {chassis_no}"
        self.success_logger.info(msg)
        self.history_logger.info(msg)

    # ── 진단 로그 (모델/단계/원인 구조화) ──────────────────────────────────────

    def log_diagnostic(self, row: int, model: str, step: str, cause: str, level: str = 'ERROR'):
        """구조화된 진단 로그.

        Args:
            row:   스프레드시트 행 번호
            model: 차량 모델명 (car_type)
            step:  에러 발생 단계 (예: '제조사_선택', '폼_필드_대기', 'listing_ID_추출')
            cause: 에러 원인 설명
            level: 'ERROR' | 'WARN' | 'INFO'
        """
        msg = f"Row {row} | 모델: {model} | 단계: {step} | 원인: {cause}"
        if level == 'ERROR':
            self.diagnostic_logger.error(msg)
            self.error_logger.error(msg)
        elif level == 'WARN':
            self.diagnostic_logger.warning(msg)
        else:
            self.diagnostic_logger.info(msg)

    def log_upload_result(self, row: int, model: str, success: bool,
                          step: str = '', cause: str = '', listing_id: str = ''):
        """업로드 성공/실패를 진단 포맷으로 기록.

        Args:
            row:        스프레드시트 행 번호
            model:      차량 모델명
            success:    성공 여부
            step:       실패 단계 (실패 시)
            cause:      실패 원인 (실패 시)
            listing_id: 비포워드 listing ID (성공 시)
        """
        if success:
            msg = f"Row {row} | 모델: {model} | 단계: 업로드_완료 | listing_id: {listing_id}"
            self.diagnostic_logger.info(msg)
            self.success_logger.info(msg)
            self.history_logger.info(msg)
        else:
            msg = f"Row {row} | 모델: {model} | 단계: {step or '알수없음'} | 원인: {cause or '알수없음'}"
            self.diagnostic_logger.error(msg)
            self.error_logger.error(msg)
            self.history_logger.error(msg)

    def _write_upload_excel_row(self, row: int, model: str, vin: str, success: bool,
                                reason: str, listing_id: str = '') -> None:
        """업로드 결과(성공/실패)를 일자별 엑셀 파일에 한 행 기록.

        파일: C:\\SynologyDrive\\...\\비포워드 에러 로그\\YYYYMMDD_업로드로그.xlsx
        컬럼: 시간 | 행번호 | 모델명 | 차대번호 | 결과 | 비고
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment

            EXCEL_LOG_DIR.mkdir(parents=True, exist_ok=True)
            today = datetime.now().strftime('%Y%m%d')
            fpath = EXCEL_LOG_DIR / f"{today}_업로드로그.xlsx"

            HEADERS = ['시간', '행번호', '모델명', '차대번호', '결과', '비고']

            if fpath.exists():
                wb = openpyxl.load_workbook(fpath)
                ws = wb.active
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = '업로드 로그'

                for col, header in enumerate(HEADERS, start=1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = PatternFill(fill_type='solid', fgColor='2C3E50')
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                ws.column_dimensions['A'].width = 20  # 시간
                ws.column_dimensions['B'].width = 8   # 행번호
                ws.column_dimensions['C'].width = 35  # 모델명
                ws.column_dimensions['D'].width = 22  # 차대번호
                ws.column_dimensions['E'].width = 10  # 결과
                ws.column_dimensions['F'].width = 55  # 비고

            next_row = ws.max_row + 1
            ws.cell(row=next_row, column=1, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=next_row, column=2, value=row)
            ws.cell(row=next_row, column=3, value=model or '')
            ws.cell(row=next_row, column=4, value=vin or '')

            result_cell = ws.cell(row=next_row, column=5)
            if success:
                result_cell.value = '성공'
                result_cell.font = Font(bold=True, color='1E8449')
                note = f"비포워드 ID: {listing_id}" if listing_id else '등록 완료'
            else:
                result_cell.value = '실패'
                result_cell.font = Font(bold=True, color='C0392B')
                note = reason or ''

            ws.cell(row=next_row, column=6, value=note)
            wb.save(fpath)

        except Exception as e:
            print(f"  [경고] 엑셀 로그 저장 실패: {e}")

    def log_upload_fail_excel(self, row: int, model: str, vin: str, reason: str) -> None:
        self._write_upload_excel_row(row, model, vin, success=False, reason=reason)

    def log_upload_success_excel(self, row: int, model: str, vin: str, listing_id: str = '') -> None:
        self._write_upload_excel_row(row, model, vin, success=True, reason='', listing_id=listing_id)
