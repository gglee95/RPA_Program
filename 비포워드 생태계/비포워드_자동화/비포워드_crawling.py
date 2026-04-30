"""
ByForward 매물 등록 크롤러 (nodriver/Chrome 기반)
차량 정보를 비포워드 외부 벤더 포털에 자동 업로드
"""
import asyncio
import json
import time
import re
import os
import sys
import shutil
from difflib import SequenceMatcher

from selenium import webdriver
from selenium.webdriver.common.by import By as SeleniumBy
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import NoSuchElementException

from config import BEFORWARD_LOGIN_URL, BEFORWARD_USERNAME, BEFORWARD_PASSWORD


class _SeleniumElementAdapter:
    def __init__(self, element):
        self._element = element

    async def click(self):
        self._element.click()

    async def send_file(self, file_path: str):
        self._element.send_keys(file_path)

    async def send_keys(self, text: str):
        self._element.send_keys(text)

    async def scroll_into_view(self):
        pass

    async def query_selector(self, css_selector: str):
        try:
            found = self._element.find_element(SeleniumBy.CSS_SELECTOR, css_selector)
            return _SeleniumElementAdapter(found)
        except NoSuchElementException:
            return None


class _SeleniumTabAdapter:
    def __init__(self, driver):
        self._driver = driver

    @property
    def url(self):
        return self._driver.current_url

    async def get(self, url: str):
        self._driver.get(url)

    async def evaluate(self, script: str):
        return self._driver.execute_script(script)

    async def select(self, css_selector: str):
        try:
            element = self._driver.find_element(SeleniumBy.CSS_SELECTOR, css_selector)
            return _SeleniumElementAdapter(element)
        except NoSuchElementException:
            return None

    async def select_all(self, css_selector: str):
        elements = self._driver.find_elements(SeleniumBy.CSS_SELECTOR, css_selector)
        return [_SeleniumElementAdapter(el) for el in elements]

    async def xpath(self, xpath: str):
        elements = self._driver.find_elements(SeleniumBy.XPATH, xpath)
        return [_SeleniumElementAdapter(el) for el in elements]

# ── 로컬 By 상수 (selenium 없이 사용) ────────────────────────────────────────
class By:
    NAME = 'name'
    ID = 'id'
    XPATH = 'xpath'
    CSS_SELECTOR = 'css'
    CLASS_NAME = 'class'

# 한국어 → 비포워드 연료 타입 매핑 (드롭다운 라벨 기준)
FUEL_MAP = {
    '가솔린': '휘발유', '휘발유': '휘발유', 'gasoline': '휘발유', 'petrol': '휘발유',
    '경유': '경유', '디젤': '경유', 'diesel': '경유',
    '가스': 'LPG', 'lpg': 'LPG', 'LPG': 'LPG',
    '전기': '전기', 'electric': '전기', 'ev': '전기',
    '하이브리드(휘발유)': '하이브리드(휘발유)', '하이브리드(가솔린)': '하이브리드(휘발유)',
    '하이브리드(경유)': '하이브리드(경유)', '하이브리드(디젤)': '하이브리드(경유)',
    '하이브리드': '하이브리드(휘발유)', 'hybrid': '하이브리드(휘발유)',
    'hybrid(lpg)': 'Hybrid(LPG)', 'Hybrid(LPG)': 'Hybrid(LPG)',
    'cng': 'CNG', 'CNG': 'CNG',
    'diesel+lpg': 'Diesel+LPG', 'Diesel+LPG': 'Diesel+LPG',
    'petrol+lpg': 'Petrol+LPG', 'Petrol+LPG': 'Petrol+LPG',
    'hydrogen': 'Hydrogen', '수소': 'Hydrogen',
}

# 한국어 → 비포워드 변속기 매핑 (드롭다운 라벨 기준)
TRANSMISSION_MAP = {
    '자동': 'Automatic', '오토': 'Automatic', 'automatic': 'Automatic',
    '수동': 'Manual', '매뉴얼': 'Manual', 'manual': 'Manual',
    'cvt': 'CVT',
    'dct': 'DCT',
    '세미오토': 'Semi Automatic', '반자동': 'Semi Automatic', 'semi': 'Semi Automatic',
    'sport at': 'Sport AT', 'sportat': 'Sport AT',
    '불명확': '불명확', '알수없음': '불명확', '불명': '불명확',
}

# 한국어 → 비포워드 제조사 매핑
MAKE_MAP = {
    '현대': 'HYUNDAI', '기아': 'KIA',
    '쌍용': 'SSANGYONG', '르노삼성': 'RENAULT SAMSUNG',
    '르노': 'RENAULT SAMSUNG', '시보레': 'CHEVROLET',
    'GM대우': 'CHEVROLET', '대우': 'DAEWOO',
    '제네시스': 'GENESIS', '삼성': 'SAMSUNG',
    '볼보': 'VOLVO', '벤츠': 'MERCEDES BENZ',
    'BMW': 'BMW', '아우디': 'AUDI',
    '폭스바겐': 'VOLKSWAGEN', '도요타': 'TOYOTA',
    '혼다': 'HONDA', '닛산': 'NISSAN',
    '푸조': 'PEUGEOT', '시트로엥': 'CITROEN',
    '랜드로버': 'LAND ROVER', '재규어': 'JAGUAR',
    '포르쉐': 'PORSCHE', '마세라티': 'MASERATI',
    '페라리': 'FERRARI', '피아트': 'FIAT',
    '알파로메오': 'ALFA ROMEO', '포드': 'FORD',
    '링컨': 'LINCOLN', '테슬라': 'TESLA',
    '렉서스': 'LEXUS', '인피니티': 'INFINITI',
    '스바루': 'SUBARU', '마쓰다': 'MAZDA',
    '미쓰비시': 'MITSUBISHI', '스즈키': 'SUZUKI',
    '크라이슬러': 'CHRYSLER', '지프': 'JEEP',
    '캐딜락': 'CADILLAC', '미니': 'MINI',
    '벤틀리': 'BENTLEY', '롤스로이스': 'ROLLS ROYCE',
}

# 차종명 내 모델 키워드 → 제조사
MODEL_KEYWORD_TO_MAKE = {
    '아반떼': 'HYUNDAI', '소나타': 'HYUNDAI', '그랜저': 'HYUNDAI',
    '투싼': 'HYUNDAI', '싼타페': 'HYUNDAI', '팰리세이드': 'HYUNDAI',
    '코나': 'HYUNDAI', '베라크루즈': 'HYUNDAI', '아이오닉': 'HYUNDAI',
    '스타렉스': 'HYUNDAI', '스타리아': 'HYUNDAI', '포터': 'HYUNDAI',
    '넥쏘': 'HYUNDAI', '캐스퍼': 'HYUNDAI', '아이30': 'HYUNDAI',
    '스포티지': 'KIA', '소렌토': 'KIA', '카니발': 'KIA',
    '셀토스': 'KIA', '스토닉': 'KIA', '모닝': 'KIA', '레이': 'KIA',
    'K3': 'KIA', 'K5': 'KIA', 'K7': 'KIA', 'K8': 'KIA', 'K9': 'KIA',
    '봉고': 'KIA', '니로': 'KIA', 'EV6': 'KIA', '카렌스': 'KIA',
    '티볼리': 'SSANGYONG', '렉스턴': 'SSANGYONG', '코란도': 'SSANGYONG',
    '무쏘': 'SSANGYONG', '액티언': 'SSANGYONG',
    'SM3': 'RENAULT SAMSUNG', 'SM5': 'RENAULT SAMSUNG',
    'SM6': 'RENAULT SAMSUNG', 'SM7': 'RENAULT SAMSUNG',
    'QM3': 'RENAULT SAMSUNG', 'QM5': 'RENAULT SAMSUNG', 'QM6': 'RENAULT SAMSUNG',
    '말리부': 'CHEVROLET', '스파크': 'CHEVROLET', '이쿼녹스': 'CHEVROLET',
    '트레일블레이저': 'CHEVROLET', '캡티바': 'CHEVROLET', '크루즈': 'CHEVROLET',
    'G70': 'GENESIS', 'G80': 'GENESIS', 'G90': 'GENESIS',
    'GV70': 'GENESIS', 'GV80': 'GENESIS',
    'E-클래스': 'MERCEDES BENZ', 'E클래스': 'MERCEDES BENZ',
    'C-클래스': 'MERCEDES BENZ', 'C클래스': 'MERCEDES BENZ',
    'S-클래스': 'MERCEDES BENZ', 'S클래스': 'MERCEDES BENZ',
    'GLC': 'MERCEDES BENZ', 'GLE': 'MERCEDES BENZ', 'GLA': 'MERCEDES BENZ',
    '3시리즈': 'BMW', '5시리즈': 'BMW', '7시리즈': 'BMW',
    'X3': 'BMW', 'X5': 'BMW', 'X1': 'BMW', 'X6': 'BMW', 'X7': 'BMW',
    'A3': 'AUDI', 'A4': 'AUDI', 'A5': 'AUDI', 'A6': 'AUDI', 'A7': 'AUDI', 'A8': 'AUDI',
    'Q3': 'AUDI', 'Q5': 'AUDI', 'Q7': 'AUDI', 'Q8': 'AUDI',
    '티구안': 'VOLKSWAGEN', '골프': 'VOLKSWAGEN', '파사트': 'VOLKSWAGEN',
    '폴로': 'VOLKSWAGEN', '투아렉': 'VOLKSWAGEN', '아테온': 'VOLKSWAGEN',
    '508': 'PEUGEOT', '308': 'PEUGEOT', '3008': 'PEUGEOT', '5008': 'PEUGEOT',
    '디스커버리': 'LAND ROVER', '레인지로버': 'LAND ROVER', '이보크': 'LAND ROVER',
    '디펜더': 'LAND ROVER', 'XE': 'JAGUAR', 'XF': 'JAGUAR',
    'XC40': 'VOLVO', 'XC60': 'VOLVO', 'XC90': 'VOLVO', 'S60': 'VOLVO', 'S90': 'VOLVO',
    '카이엔': 'PORSCHE', '마칸': 'PORSCHE', '파나메라': 'PORSCHE',
    '모델3': 'TESLA', '모델Y': 'TESLA', '모델S': 'TESLA', '모델X': 'TESLA',
    'ES': 'LEXUS', 'RX': 'LEXUS', 'NX': 'LEXUS', 'IS': 'LEXUS',
    'UX': 'LEXUS', 'GX': 'LEXUS', 'LX': 'LEXUS', 'LS': 'LEXUS',
    'LC': 'LEXUS', 'GS': 'LEXUS', 'RC': 'LEXUS', 'CT': 'LEXUS',
    '렉서스': 'LEXUS', 'UX250': 'LEXUS', 'UX200': 'LEXUS',
}

WMI_TO_MAKE = {
    'KMH': 'HYUNDAI', 'KMF': 'HYUNDAI', 'KME': 'HYUNDAI',
    'KNA': 'KIA', 'KNB': 'KIA', 'KNC': 'KIA', 'KND': 'KIA', 'KNM': 'KIA',
    'KPT': 'SSANGYONG', 'KPB': 'SSANGYONG',
    'KL1': 'CHEVROLET', 'KL8': 'CHEVROLET',
    'KLA': 'RENAULT SAMSUNG', 'KLY': 'RENAULT SAMSUNG',
    'KMT': 'GENESIS', 'KMX': 'GENESIS',
    'VF1': 'RENAULT SAMSUNG',
    'WBA': 'BMW', 'WBS': 'BMW', 'WBY': 'BMW', 'WBW': 'BMW',
    'WDD': 'MERCEDES BENZ', 'WDC': 'MERCEDES BENZ', 'W1N': 'MERCEDES BENZ',
    'W1K': 'MERCEDES BENZ', 'W1V': 'MERCEDES BENZ',
    'WAU': 'AUDI', 'WAP': 'PORSCHE', 'WP0': 'PORSCHE', 'WP1': 'PORSCHE',
    'WVW': 'VOLKSWAGEN', 'WVG': 'VOLKSWAGEN', 'WV1': 'VOLKSWAGEN', 'WV2': 'VOLKSWAGEN',
    'VF3': 'PEUGEOT', 'VR1': 'PEUGEOT', 'VR3': 'PEUGEOT',
    'YV1': 'VOLVO', 'YV4': 'VOLVO',
    'SAL': 'LAND ROVER', 'SAJ': 'JAGUAR',
    'ZAR': 'ALFA ROMEO', 'ZFA': 'FIAT', 'ZFF': 'FERRARI', 'ZAM': 'MASERATI',
    'WF0': 'FORD', 'WF1': 'FORD', 'WDB': 'MERCEDES BENZ',
    'JTD': 'TOYOTA', 'JTE': 'TOYOTA', 'JTM': 'TOYOTA', 'JTJ': 'TOYOTA',
    'JTH': 'LEXUS', 'JT8': 'LEXUS', 'JT3': 'LEXUS', 'JT4': 'LEXUS', 'JT6': 'LEXUS',
    'JHM': 'HONDA', 'JN1': 'NISSAN', 'JN8': 'NISSAN',
    'JF1': 'SUBARU', 'JF2': 'SUBARU',
    'JMZ': 'MAZDA', 'JM1': 'MAZDA',
    'JS1': 'SUZUKI', 'JS2': 'SUZUKI',
    'MLH': 'HONDA',
    '1G1': 'CHEVROLET', '1GC': 'CHEVROLET', '2G1': 'CHEVROLET',
    '1FA': 'FORD', '1FM': 'FORD', '3FA': 'FORD',
    '1C4': 'CHRYSLER', '1C6': 'DODGE',
    '5YJ': 'TESLA',
}

COLOR_MAP_KO = {
    # 화이트 계열
    '흰색': '화이트', '백색': '화이트', '진주흰색': '화이트', '순백색': '화이트',
    '화이트': '화이트', '크림화이트': '화이트', '크리스탈화이트': '화이트', '아이보리': '화이트',
    '로얄화이트': '화이트', '플래티넘화이트': '화이트', '세레니티화이트': '화이트',
    '아틱화이트': '화이트', '스노우화이트': '화이트', '오팔화이트': '화이트',
    'white': '화이트',
    # 블랙 계열
    '검정': '블랙', '검은색': '블랙', '블랙': '블랙', '제트블랙': '블랙', '팬텀블랙': '블랙',
    '어비스블랙': '블랙', '오닉스블랙': '블랙', '나이트블랙': '블랙', '딥블랙': '블랙',
    '크리스탈블랙': '블랙',
    'black': '블랙',
    # 실버 계열
    '은색': '실버', '은회색': '실버', '실버': '실버', '문라이트실버': '실버', '그라파이트실버': '실버',
    '이레나실버': '실버', '플루이드메탈': '실버', '알루미늄실버': '실버', '스틸실버': '실버',
    '문라이트': '실버', '플레티넘실버': '실버',
    'silver': '실버',
    # 그레이 계열
    '회색': '그레이', '쥐색': '그레이', '그레이': '그레이', '다크그레이': '그레이', '차콜그레이': '그레이', '쉐도우그레이': '그레이',
    '어반그레이': '그레이', '팬텀그레이': '그레이', '스틸그레이': '그레이', '미스트그레이': '그레이',
    '스모키그레이': '그레이', '틴더그레이': '그레이', '차콜': '그레이',
    'gray': '그레이', 'grey': '그레이',
    # 레드 계열
    '빨간색': '레드', '적색': '레드', '레드': '레드', '다크체리': '레드', '와인레드': '레드',
    '이그니트레드': '레드', '크림슨레드': '레드', '그라나다레드': '레드', '버건디': '레드',
    '보르도': '레드', '캔디레드': '레드', '체리레드': '레드',
    'red': '레드',
    # 블루 계열
    '파란색': '블루', '청색': '블루', '블루': '블루', '네이비블루': '블루', '다크블루': '블루',
    '세레니티블루': '블루', '아이온블루': '블루', '마그마블루': '블루', '슈퍼소닉블루': '블루',
    '라피스블루': '블루', '머젯블루': '블루', '스틸블루': '블루', '하이퍼블루': '블루',
    '하늘색': '블루', '하늘': '블루', '네이비': '블루', '인디고': '블루',
    'blue': '블루',
    # 옐로우 계열
    '노란색': '옐로우', '황색': '옐로우', '옐로우': '옐로우',
    'yellow': '옐로우',
    # 오렌지 계열
    '주황색': '오렌지', '오렌지': '오렌지',
    'orange': '오렌지',
    # 그린 계열
    '연두색': '그린', '녹색': '그린', '그린': '그린', '카키': '그린', '올리브': '그린',
    '민트': '그린', '어반카키': '그린', '에코그린': '그린',
    'green': '그린',
    # 핑크 계열
    '분홍색': '핑크', '핑크': '핑크', '마젠타': '핑크', '체리블라썸': '핑크', '로즈': '핑크',
    'pink': '핑크',
    # 브라운 계열
    '갈색': '브라운', '브라운': '브라운', '어번테라코타': '브라운',
    '코퍼': '브라운', '카파': '브라운', '모카': '브라운',
    'brown': '브라운',
    # 브론즈 (별도 항목)
    '브론즈': '브론즈', 'bronze': '브론즈',
    # 베이지 계열
    '베이지': '베이지', '샌드베이지': '베이지', '크림베이지': '베이지',
    'beige': '베이지',
    # 골드 계열
    '금색': '골드', '골드': '골드', '샴페인골드': '골드', '샴페인': '골드',
    'gold': '골드',
    # 퍼플 계열
    '보라색': '퍼플', '퍼플': '퍼플', '바이올렛': '퍼플',
    'purple': '퍼플',
    # 마룬 (별도 옵션)
    '마룬': '마룬', '와인': '마룬', '버건디레드': '마룬',
    # 펄 (별도 옵션)
    '펄': '펄',
}

BODY_TYPE_MAP = {
    'SUV': ['SUV', '싼타페', '투싼', '스포티지', '소렌토', '코나', '셀토스',
            '티볼리', '팰리세이드', '모하비', '베라크루즈', 'QM', '트레일블레이저',
            '이쿼녹스', '캐스퍼', '스토닉', '니로', '아이오닉5', 'EV6', '넥쏘',
            '렉스턴', '코란도', '액티언'],
    'Minivan': ['스타렉스', '카니발', '그랜드스타렉스', '카렌스', '올란도', '스타리아'],
    'Truck': ['봉고', '포터', '마이티', '트럭', '화물'],
    'Hatchback': ['i30', 'i20', 'i10', '프라이드', '해치', 'MINI', '폴로'],
    'Wagon': ['투어링', '왜건', 'SW'],
    'Coupe': ['쿠페', 'RC', 'GR쿠페'],
    'Van': ['밴'],
}
BODY_TYPE_TEXT_CANDIDATES = {
    'SUV': ['SUV'],
    'Minivan': ['미니 밴', '미니버스', '밴'],
    'Truck': ['트럭', '픽업'],
    'Hatchback': ['해치백', '경차'],
    'Wagon': ['웨건'],
    'Coupe': ['쿠페', '컨버터블'],
    'Van': ['밴'],
    'Sedan': ['세단'],
    'Bus': ['버스', '버스 20좌석', '미니버스'],
    'Pickup': ['픽업', '트럭'],
    'Convertible': ['컨버터블', '쿠페'],
    'Motorhome': ['Motorhome'],
}

BF_MAKE_MODEL_SHEET_ID = "1ZCI7DXHiIeRxYfQzVVJqauzgRS29xybAuAYLjbU19o4"
BF_MAKE_MODEL_WORKSHEETS = ("우선 대상 메이커", "RAW")

KO_MODEL_ALIAS = {
    '투싼': 'tucson', '싼타페': 'santafe', '아반떼': 'avante',
    '소나타': 'sonata', '그랜저': 'grandeur', '코나': 'kona',
    '팰리세이드': 'palisade', '넥쏘': 'nexo', '아이오닉': 'ioniq',
    '스타렉스': 'starex', '스타리아': 'staria', '스포티지': 'sportage',
    '쏘렌토': 'sorento', '소렌토': 'sorento', '카니발': 'carnival',
    '셀토스': 'seltos', '스토닉': 'stonic', '모닝': 'morning',
    '레이': 'ray', '니로': 'niro', '티볼리': 'tivoli',
    '렉스턴': 'rexton', '코란도': 'korando',
    '말리부': 'malibu', '스파크': 'spark', '크루즈': 'cruze', '캡티바': 'captiva',
    '이쿼녹스': 'equinox',
    'GV70': 'gv70', 'GV80': 'gv80', 'G70': 'g70', 'G80': 'g80', 'G90': 'g90',
}

MODEL_PREFIX_STOPWORDS = {'올', '올뉴', '더', '더뉴', '뉴', '신형', 'all', 'new', 'the'}
MODEL_SUFFIX_STOPWORDS = ('디젤', '가솔린', '하이브리드', '전기', 'lpg')
MODEL_SELECT_CANDIDATES = ('TempVehDetails[model_id]', 'TempVehDetails[model_code]')
BODY_TYPE_SELECT_CANDIDATES = ('TempVehDetails[type_id]', 'TempVehDetails[type_2_id]')
FINAL_CLICK_XPATHS = ('//*[@id="chk-compensation-agree"]',)
SUBMIT_BUTTON_XPATH = '/html/body/div[2]/div[1]/div/div/div/div/div/div/div/input'


# ═══════════════════════════════════════════════════════════════════════════════
# BefowordCrawler  (nodriver / Chrome 기반)
# ═══════════════════════════════════════════════════════════════════════════════

class BefowordCrawler:
    """ByForward 외부 벤더 포털 - 매물 등록 (nodriver/Chrome)"""

    WAIT_TIMEOUT = 10
    FORM_URL = "https://external-vendor.beforward.jp/tempVehDetails/edit"

    _4WD_KEYWORDS_EN = ['AWD', '4WD', 'XDRIVE', '4MATIC', '4MOTION']
    _4WD_KEYWORDS_KO = ['콰트로', '4매틱', '4모션', '4륜', 'x드라이브']

    def __init__(self, headless: bool = False):
        self.headless = headless
        self.browser = None
        self.tab = None
        self._bf_make_models = {}
        self._bf_model_ref_loaded = False
        self._vehicle_ref_table = []
        self._last_downloaded_image_files = []
        # 에러 진단용 컨텍스트 (encar_soldout_monitor에서 읽어 로깅)
        self._last_error_step = ''   # 실패 단계
        self._last_error_cause = ''  # 실패 원인
        self._listing_submitted = False  # 폼 제출(listing_id 추출) 성공 여부
        self._load_vehicle_ref_table()

    # ── 이벤트 루프 관리 ──────────────────────────────────────────────────────

    def _get_loop(self):
        try:
            loop = asyncio.get_event_loop()
        except RuntimeError:
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
        if loop.is_closed():
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
        return loop

    def _run(self, coro):
        return self._get_loop().run_until_complete(coro)

    # ── 하위 호환 속성 ────────────────────────────────────────────────────────

    @property
    def driver(self):
        """하위 호환: tab 반환"""
        return self.tab

    # ── 드라이버 초기화 ───────────────────────────────────────────────────────

    def _setup_driver(self):
        # 이전 크래시로 남은 stale chromedriver 프로세스 정리
        try:
            import subprocess
            subprocess.run(
                ['taskkill', '/F', '/IM', 'chromedriver.exe'],
                capture_output=True, timeout=5
            )
        except Exception:
            pass
        self._run(self._async_setup())

    async def _async_setup(self):
        options = Options()
        options.add_argument('--window-size=1920,1080')
        options.add_argument('--lang=ko-KR,ko')
        options.add_argument('--disable-blink-features=AutomationControlled')
        if self.headless:
            options.add_argument('--headless=new')

        driver = webdriver.Chrome(options=options)
        self.browser = driver
        self.tab = _SeleniumTabAdapter(driver)

    # ── 재원표 로드 / 검색 ────────────────────────────────────────────────────

    def _load_vehicle_ref_table(self):
        """국산차·수입차 재원표 Excel 파일의 모든 시트를 로드한다."""
        import openpyxl
        base_dir = os.path.dirname(os.path.abspath(__file__))
        for fname in ['국산차 재원표.xlsx', '수입차 재원표.xlsx']:
            fpath = os.path.join(base_dir, fname)
            try:
                wb = openpyxl.load_workbook(fpath, data_only=True)
                sheet_count = 0
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        keyword = row[1]
                        if not keyword:
                            continue
                        self._vehicle_ref_table.append({
                            'keyword': str(keyword).strip(),
                            'bf_model': row[3],
                            'bf_make':  row[4],
                            'bf_type':  row[5],
                            'doors':    row[6],
                            'seats':    row[7],
                        })
                    sheet_count += 1
            except Exception as e:
                print(f"  [경고] 재원표 로드 실패 ({fname}): {e}")

    def _lookup_vehicle_ref(self, car_type: str) -> dict | None:
        """엑셀 키워드가 car_type 안에 독립된 단어로 완전히 포함될 때만 매칭.

        예) keyword="C-클래스", car_type="GLC-클래스 300" → 앞에 'L'이 붙어있으므로 불일치
            keyword="GLC-클래스", car_type="GLC-클래스 300" → 앞뒤 경계 OK → 일치
        """
        import re
        if not car_type or not self._vehicle_ref_table:
            return None
        for entry in self._vehicle_ref_table:
            kw = re.escape(entry['keyword'])
            # 키워드 앞뒤에 한글/영문/숫자가 없어야 독립된 단어로 판단
            if re.search(r'(?<![가-힣a-zA-Z0-9])' + kw + r'(?![가-힣a-zA-Z0-9])', car_type):
                return entry
        return None

    def _map_ref_body_type(self, bf_type_ko: str) -> str:
        mapping = {
            '세단': 'Sedan', 'SUV': 'SUV', '미니밴': 'Minivan', '밴': 'Van',
            '트럭': 'Truck', '해치백': 'Hatchback', '왜건': 'Wagon',
            '쿠페': 'Coupe', '컨버터블': 'Convertible', '픽업': 'Pickup',
        }
        val = str(bf_type_ko).strip()
        return mapping.get(val, val)

    def _cleanup_downloaded_images(self, image_files: list[str]) -> None:
        for fpath in image_files:
            try:
                if os.path.isfile(fpath):
                    os.remove(fpath)
                parent = os.path.dirname(fpath)
                if os.path.isdir(parent) and not os.listdir(parent):
                    shutil.rmtree(parent, ignore_errors=True)
            except Exception as e:
                print(f"  [경고] 이미지 삭제 실패 ({os.path.basename(fpath)}): {e}")

    # ── 공개 sync 인터페이스 ──────────────────────────────────────────────────

    def login(self) -> bool:
        if not self.browser:
            self._setup_driver()
        return self._run(self._async_login())

    def fill_vehicle_data(self, car_info, auto_submit: bool = False) -> bool:
        return self._run(self._async_fill_vehicle_data(car_info, auto_submit))

    def upload_images_for_listing(self, listing_id: str, image_files: list[str]) -> bool:
        """저장된 매물의 수정 페이지에서 이미지를 업로드한다 (2단계)."""
        return self._run(self._async_upload_images_for_listing(listing_id, image_files))

    def close_popups(self):
        """팝업 닫기 (encar_soldout_monitor 호환)"""
        if self.tab:
            self._run(self._async_dismiss_popup())

    def close(self):
        if self.browser:
            try:
                self._run(self._async_close())
            except Exception:
                pass
            self.browser = None
            self.tab = None

    # ── async 내부 구현 ───────────────────────────────────────────────────────

    async def _async_close(self):
        if self.browser:
            self.browser.quit()

    def _capture_screenshot(self, label: str = "error") -> str:
        """에러 발생 시 스크린샷 캡처 → screenshots/ 폴더에 저장"""
        try:
            ss_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "screenshots")
            os.makedirs(ss_dir, exist_ok=True)
            ts = time.strftime("%Y%m%d_%H%M%S")
            safe_label = re.sub(r'[^a-zA-Z0-9가-힣_-]', '_', label)[:50]
            fname = f"{ts}_{safe_label}.png"
            fpath = os.path.join(ss_dir, fname)
            self.browser.save_screenshot(fpath)
            return fpath
        except Exception as e:
            print(f"  [경고] 스크린샷 저장 실패: {e}")
            return ""

    async def _inject_alert_interceptor(self):
        """JS alert/confirm을 가로채 window._pendingAlert에 저장"""
        try:
            await self.tab.evaluate("""
                window._pendingAlert = '';
                window.alert = function(msg) { window._pendingAlert = String(msg || ''); };
                window.confirm = function(msg) { window._pendingAlert = String(msg || ''); return true; };
                window.onbeforeunload = null;
            """)
        except Exception:
            pass

    async def _inject_duplicate_modal_auto_closer(self):
        """DOM MutationObserver로 duplicate 모달이 나타나는 즉시 OK/닫기 버튼 자동 클릭"""
        try:
            await self.tab.evaluate("""
                (function() {
                    if (window._dupModalObserver) {
                        window._dupModalObserver.disconnect();
                    }
                    function closeDupModal(root) {
                        var text = (root.textContent || '').toLowerCase();
                        if (!text.includes('duplicate') && !text.includes('can not upload')) return;
                        // OK 버튼 우선
                        var els = Array.from(root.querySelectorAll('button, a, input[type="button"]'));
                        for (var i = 0; i < els.length; i++) {
                            var t = els[i].textContent.trim().toUpperCase();
                            if (t === 'OK' || t === '×' || t === 'X') {
                                els[i].click();
                                console.log('[AUTO] duplicate modal closed: ' + t);
                                return;
                            }
                        }
                        // fallback: 아무 버튼이나 클릭
                        for (var i = 0; i < els.length; i++) {
                            if (els[i].offsetParent !== null) {
                                els[i].click();
                                console.log('[AUTO] duplicate modal closed (fallback)');
                                return;
                            }
                        }
                    }
                    window._dupModalObserver = new MutationObserver(function(mutations) {
                        mutations.forEach(function(m) {
                            m.addedNodes.forEach(function(node) {
                                if (node.nodeType === 1) closeDupModal(node);
                            });
                        });
                    });
                    window._dupModalObserver.observe(document.body, {childList: true, subtree: true});
                })();
            """)
        except Exception:
            pass

    async def _async_login(self) -> bool:
        try:
            await self.tab.get(BEFORWARD_LOGIN_URL)
            await asyncio.sleep(1)

            if 'login' not in self.tab.url.lower():
                return True

            await self.tab.evaluate(f"""
                (function() {{
                    var e = document.querySelector('input[name="data[VendorUser][email]"]');
                    var p = document.querySelector('input[name="data[VendorUser][password]"]');
                    if (e) e.value = {json.dumps(BEFORWARD_USERNAME)};
                    if (p) p.value = {json.dumps(BEFORWARD_PASSWORD)};
                }})()
            """)
            await asyncio.sleep(0.2)

            try:
                submit = await self.tab.select('button[type="submit"]')
                await submit.click()
            except Exception:
                await self.tab.evaluate("document.querySelector('button[type=\"submit\"]').click()")
            await asyncio.sleep(2)

            if 'login' in self.tab.url.lower():
                print("[오류] 로그인 실패")
                return False

            return True
        except Exception as e:
            print(f"[오류] 로그인 중 오류: {e}")
            import traceback
            traceback.print_exc()
            return False

    # ── 페이지 이동 (버튼 클릭 기반) ────────────────────────────────────────────

    async def _navigate_to_new_form(self) -> bool:
        """신규 등록 폼으로 이동 (직접 URL 이동 - 가장 안정적)"""
        await self.tab.get(self.FORM_URL)
        # #bulk_confirm_form 요소가 나타날 때까지 대기 (최대 10초)
        for _ in range(20):
            await asyncio.sleep(0.5)
            if 'edit' not in self.tab.url.lower():
                continue
            found = await self.tab.evaluate(
                "return !!document.querySelector('#bulk_confirm_form')"
            )
            if found:
                return True
        return False

    async def _navigate_to_edit_page(self, listing_id: str) -> bool:
        """listing ID를 알고 있으므로 edit URL에 직접 이동"""
        edit_url = f"https://external-vendor.beforward.jp/tempVehDetails/edit/{listing_id}"
        await self.tab.get(edit_url)
        await asyncio.sleep(2)
        cur = self.tab.url
        if f'/edit/{listing_id}' in cur.lower():
            return True
        # 리디렉션됐으면 로그인 페이지일 수 있음 - 재로그인 후 재시도
        if 'login' in cur.lower():
            print(f"  [경고] 세션 만료 - 재로그인")
            if self.login():
                await self.tab.get(edit_url)
                await asyncio.sleep(2)
                cur = self.tab.url
                if f'/edit/{listing_id}' in cur.lower():
                    return True
        print(f"  [경고] 수정 페이지 이동 실패: {cur}")
        self._capture_screenshot(f"edit_nav_direct_failed_{listing_id}")
        return False

    async def _click_image_tab_from_edit(self) -> bool:
        """수정 페이지에서 이미지 탭 버튼 클릭"""
        IMAGE_TAB_XPATHS = [
            '/html/body/div[2]/div[1]/div[1]/div/div/div[1]/div[1]/div[2]/button',
            '/html/body/div[2]/div[1]/div[1]/div/div/div[1]/div[1]/div[2]/a',
            "//button[contains(text(), '画像') or contains(text(), 'Image') or contains(text(), '이미지') or contains(text(), 'Photo')]",
            "//a[contains(text(), '画像') or contains(text(), 'Image') or contains(text(), '이미지') or contains(text(), 'Photo')]",
            "//a[contains(@href, 'photo/upload')]",
        ]
        IMAGE_TAB_CSS = [
            'a[href*="photo/upload"]',
            '.nav-tabs a:nth-child(2)',
            'button.btn-info',
        ]

        for css in IMAGE_TAB_CSS:
            try:
                elem = await self.tab.select(css)
                if elem:
                    await elem.click()
                    await asyncio.sleep(0.5)
                    return True
            except Exception:
                pass

        for xpath in IMAGE_TAB_XPATHS:
            if await self._click_xpath(xpath, "이미지 탭", timeout=5):
                await asyncio.sleep(0.5)
                return True

        print(f"  [경고] 이미지 탭 버튼 찾기 실패")
        self._capture_screenshot("image_tab_not_found")
        return False

    async def _click_condition_tab_from_current(self) -> bool:
        """현재 페이지에서 성능점검 탭 버튼 클릭"""
        COND_XPATHS = [
            '//*[@id="bulk_confirm_form"]/div/a[2]',
            "//a[contains(@href, 'ConditionsSheet')]",
            "//a[contains(text(), 'Condition') or contains(text(), '성능') or contains(text(), '점검')]",
            "//button[contains(text(), 'Condition') or contains(text(), '성능')]",
        ]
        COND_CSS = [
            'a[href*="ConditionsSheet"]',
            'a[href*="conditionssheet"]',
        ]

        for css in COND_CSS:
            try:
                elem = await self.tab.select(css)
                if elem:
                    await elem.click()
                    await asyncio.sleep(0.5)
                    return True
            except Exception:
                pass

        for xpath in COND_XPATHS:
            if await self._click_xpath(xpath, "성능점검 탭", timeout=5):
                await asyncio.sleep(0.5)
                return True

        print(f"  [경고] 성능점검 탭 버튼 찾기 실패")
        self._capture_screenshot("condition_tab_not_found")
        return False

    # ── 메인 폼 입력 ──────────────────────────────────────────────────────────

    async def _async_fill_vehicle_data(self, car_info, auto_submit: bool = False) -> bool:
        # 에러 컨텍스트 초기화
        self._last_error_step = ''
        self._last_error_cause = ''
        self._listing_submitted = False
        try:

            # 저장 전 기존 listing ID 수집 생략 (관리 페이지 로드 → Chrome OOM 크래시 원인)
            # → 저장 후 URL에서 직접 ID 추출
            pre_save_ids = set()

            # 신규 등록 폼으로 이동
            await self._navigate_to_new_form()
            await self._inject_alert_interceptor()

            # 0. 재원표 조회 - 재원표에 없으면 업로드 불가
            ref = self._lookup_vehicle_ref(car_info.car_type or '')
            if not ref or not ref.get('bf_model'):
                self._last_error_step = '재원표_조회'
                self._last_error_cause = f"재원표에 모델 없음: '{car_info.car_type}' - 엑셀 재원표에 해당 차종을 추가하세요"
                print(f"  [오류] 재원표 미등록 차종: '{car_info.car_type}'")
                return False

            # 1. 제조사+모델 선택 (Select2 클릭 직후 model options 로드, 대기 시 초기화됨)
            make_name = ref['bf_make']
            bf_model = str(ref['bf_model']).strip()
            if make_name:
                await self._select_make(make_name, car_info.car_type, bf_model)
                # make 선택 후 model_id 옵션 로드 대기 (최대 5초)
                # ※ model_id의 change 이벤트는 발생시키지 않음 → 2차 AJAX/DOM 교체 없음
                await asyncio.sleep(1.0)

            # 4. 모델년도
            year = car_info.year_month or ""
            if year:
                await self._js_select_by_value('TempVehDetails[registration_year]', year, "모델년도")

            # 5. 제조년월
            if year:
                await self._js_select_by_value('TempVehDetails[manufacture_year]', year, "제조년월")

            # 6. CBM 기본값 설정 (차량 타입별, 단위: cm)
            # ※ mileage 클릭은 제거 - AJAX가 DOM 전체를 교체하여 다른 필드들이 사라짐
            cbm_defaults = {
                'SUV': (470, 190, 175),
                'Minivan': (500, 195, 185),
                'Truck': (500, 180, 180),
                'Hatchback': (420, 175, 150),
                'Wagon': (470, 180, 155),
                'Coupe': (450, 180, 140),
                'Van': (500, 185, 195),
                'Sedan': (470, 180, 150),
            }
            _bt = self._get_body_type(car_info.car_type)
            dl, dw, dh = cbm_defaults.get(_bt, (470, 180, 150))
            m3 = round(dl * dw * dh / 1000000, 3)
            await self._fill_text_by_name('TempVehDetails[length]', str(dl), "길이(cm)")
            await self._fill_text_by_name('TempVehDetails[width]', str(dw), "너비(cm)")
            await self._fill_text_by_name('TempVehDetails[height]', str(dh), "높이(cm)")
            await self._fill_text_by_name('TempVehDetails[m3]', str(m3), "M3")

            # 7-1. 차대번호
            await self._fill_text_by_name('TempVehDetails[chassis_no]',
                                          car_info.inspection_chassis_no or "", "차대번호")

            # 8. 차량 타입
            if ref and ref.get('bf_type'):
                body_type = self._map_ref_body_type(ref['bf_type'])
            else:
                body_type = self._get_body_type(car_info.car_type)
            await self._select_body_type(body_type)

            # 9. 주행거리
            await self._fill_text_by_name('TempVehDetails[mileage]',
                                          car_info.mileage or "", "주행거리")

            # 10. 배기량
            await self._fill_text_by_name('TempVehDetails[engine_capacity]',
                                          car_info.displacement or "", "배기량")

            # 11. 연료 타입 - CSS 선택자 우선, name 속성 폴백
            _FUEL_CSS = '#bulk_confirm_form div > div > div:first-child table:nth-of-type(3) tr:nth-child(6) select'
            fuel_bf = self._map_fuel(car_info.fuel_type)
            if fuel_bf:
                fuel_opts = await self._css_get_select_options(_FUEL_CSS)
                if not fuel_opts:
                    fuel_opts = await self._js_get_select_options('TempVehDetails[fuel_id]')
                fuel_value = next(
                    (o['value'] for o in fuel_opts if o['text'].strip() == fuel_bf), None
                )
                if fuel_value:
                    ok = await self._css_select_by_value(_FUEL_CSS, fuel_value, "연료")
                    if not ok:
                        await self._js_select_by_value('TempVehDetails[fuel_id]', fuel_value, "연료")

            # 12. 변속기 - CSS 선택자 우선, name 속성 폴백
            _TRANS_CSS = '#bulk_confirm_form div > div > div:first-child table:nth-of-type(3) tr:nth-child(8) select'
            trans_bf = self._map_transmission(car_info.transmission)
            if trans_bf:
                trans_opts = await self._css_get_select_options(_TRANS_CSS)
                if not trans_opts:
                    trans_opts = await self._js_get_select_options('TempVehDetails[transmission_id]')
                trans_value = next(
                    (o['value'] for o in trans_opts if o['text'].strip() == trans_bf), None
                )
                if trans_value:
                    ok = await self._css_select_by_value(_TRANS_CSS, trans_value, "변속기")
                    if not ok:
                        await self._js_select_by_value('TempVehDetails[transmission_id]', trans_value, "변속기")

            # 13. 핸들 - 좌핸들 고정 (name 방식 + XPath label 클릭 병행)
            await self._click_radio_by_name('TempVehDetails[steering]', '2', "핸들(좌핸들)")
            await self._click_xpath(
                '//*[@id="bulk_confirm_form"]/div/div/div[2]/table[5]/tbody/tr[2]/td[3]/label',
                "핸들 라디오 label"
            )

            # 14. 문 개수
            doors_val = str(int(ref['doors'])) if ref and ref.get('doors') else '4'
            await self._fill_or_select_by_name('TempVehDetails[doors]', doors_val, "문개수")

            # 15. 구동 방식
            drive_val = self._get_drive_type(car_info.car_type)
            if drive_val:
                await self._click_radio_by_name('TempVehDetails[drive_type]', drive_val, "구동방식")

            # 16. 색상 - CSS 선택자 우선, name 속성 폴백
            _COLOR_CSS = '#bulk_confirm_form div > div > div:first-child table:nth-of-type(3) tr:nth-child(10) select'
            if car_info.color:
                mapped_color = self._map_color(car_info.color)
                color_opts = await self._css_get_select_options(_COLOR_CSS)
                if not color_opts:
                    color_opts = await self._js_get_select_options('TempVehDetails[ext_color_id]')
                color_value = next(
                    (o['value'] for o in color_opts if o['text'].strip() == mapped_color), None
                )
                if not color_value:
                    color_value = next(
                        (o['value'] for o in color_opts if o['text'].strip() == '기타'), None
                    )
                if color_value:
                    ok = await self._css_select_by_value(_COLOR_CSS, color_value, "색상")
                    if not ok:
                        await self._js_select_by_value('TempVehDetails[ext_color_id]', color_value, "색상")

            # 17. 좌석수
            if ref and ref.get('seats'):
                await self._fill_text_by_name('TempVehDetails[seats]',
                                              str(int(ref['seats'])), "좌석수")
            elif car_info.seating_capacity:
                seats = ''.join(filter(str.isdigit, car_info.seating_capacity))
                if seats:
                    await self._fill_text_by_name('TempVehDetails[seats]', seats, "좌석수")

            # 18. 가격 (ID 또는 name 방식 시도)
            price_ok = await self._fill_text_by_id('trade-price-input', car_info.price or "", "가격")
            if not price_ok:
                await self._fill_text_by_name('TempVehDetails[trade_price_input]', car_info.price or "", "가격(name)")
            # 할인가 = 거래 금액보다 낮게 (0으로 설정)
            await self._fill_text_by_name('TempVehDetails[available_discount]', '0', "할인가")

            # 19. 재고 위치
            await self._js_select_by_text('TempVehDetails[stock_place_id]', 'KOREA', "재고위치")

            # 20. 옵션 체크박스
            MANDATORY_OPTIONS = ['에어컨', '파워핸들', '파워 윈도우', '에어백', 'ABS', 'AM/FM 라디오']
            options = list(car_info.options) if hasattr(car_info, 'options') and car_info.options else []
            existing_mapped = {o.mapped_name for o in options}
            from 엔카_crawling import OptionItem
            for opt_name in MANDATORY_OPTIONS:
                if opt_name not in existing_mapped:
                    options.append(OptionItem(name=opt_name, mapped_name=opt_name))
            if options:
                await self._fill_options(options)

            # 21. 이미지 다운로드 (Google Drive 또는 망고카 링크)
            drive_link = getattr(car_info, 'drive_link', '') or ''
            sheet_row = getattr(car_info, 'sheet_row', 0) or 0
            image_files = []
            is_gdrive = drive_link and ('drive.google.com' in drive_link or 'docs.google.com' in drive_link)
            is_mango = drive_link and 'mangoworldcar.com' in drive_link
            if is_gdrive:
                image_files = self._download_images_from_drive_link(drive_link, sheet_row)
                self._last_downloaded_image_files = list(image_files)
                if not image_files:
                    print(f"  [경고] 구글드라이브 이미지 다운로드 실패 - 이미지 없이 등록 진행")
            elif is_mango:
                image_files = self._download_images_from_mango_link(drive_link, sheet_row)
                self._last_downloaded_image_files = list(image_files)
                if not image_files:
                    print(f"  [경고] 망고카 이미지 다운로드 실패 - 이미지 없이 등록 진행")
            elif drive_link:
                print(f"  [경고] 알 수 없는 링크 형식 → 이미지 없이 진행: {drive_link[:60]}")
            else:
                print(f"  [경고] 드라이브 링크 없음 - 이미지 없이 진행")

            # 22. 필수 클릭
            await self._click_required_xpaths()

            # 22-1. 4WD 옵션 선택 (4WD 차량만 tr[8] 클릭)
            # 주의: 기존 tr[2] 항상 클릭은 AM/FM 라디오(이미 기본 체크됨)를 토글로 풀어버려 제거
            _is_4wd = self._get_drive_type(car_info.car_type or '') == '3'
            if _is_4wd:
                await self._click_option_radio(8)

            # 23. 저장 버튼
            if auto_submit:
                try:
                    print("  [INFO] submit listing form")
                    submitted = await self.tab.evaluate(f"""
                        (function() {{
                            var btn = document.evaluate({json.dumps(SUBMIT_BUTTON_XPATH)},
                                document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue;
                            if (!btn) return false;
                            btn.removeAttribute('disabled');
                            btn.scrollIntoView({{block:'center'}});
                            btn.click();
                            return true;
                        }})()
                    """)
                    if not submitted:
                        submitted = await self.tab.evaluate("""
                            (function() {
                                var selectors = [
                                    'button[type="submit"]',
                                    'input[type="submit"]',
                                    'input[name="edit"]',
                                    'button[name="edit"]',
                                    'input[value*="Save"]',
                                    'button:enabled'
                                ];
                                for (var i = 0; i < selectors.length; i++) {
                                    var btn = document.querySelector(selectors[i]);
                                    if (!btn) continue;
                                    btn.removeAttribute('disabled');
                                    btn.scrollIntoView({block:'center'});
                                    btn.click();
                                    return true;
                                }
                                return false;
                            })()
                        """)
                    if not submitted:
                        print("  [오류] 저장 버튼을 찾지 못함")
                        self._capture_screenshot("submit_button_not_found")
                        self._last_error_step = '폼_저장_버튼'
                        self._last_error_cause = 'submit button not found'
                        return False

                    # 저장 후 리디렉션 대기 (최대 10초)
                    for _w in range(30):
                        await asyncio.sleep(1)
                        cur = self.tab.url
                        if 'edit/' in cur.lower() or 'photo/upload' in cur.lower():
                            break
                    print(f"  [INFO] after submit url: {self.tab.url}")
                    # interceptor로 캡처된 alert 확인
                    _alerts = await self.tab.evaluate("JSON.stringify(window._allAlerts || window._pendingAlert || '')")

                    # validation 에러 체크
                    errors = await self.tab.evaluate("""
                        (function() {
                            var msgs = [];
                            document.querySelectorAll('.error, .text-danger, .error-message, [class*="error"]').forEach(function(el) {
                                if (el.offsetParent !== null && el.textContent.trim())
                                    msgs.push(el.textContent.trim().substring(0, 100));
                            });
                            return JSON.stringify(msgs);
                        })()
                    """)
                    try:
                        err_list = json.loads(str(errors)) if errors else []
                    except Exception:
                        err_list = []
                    if err_list:
                        # "이미 등록된 차대번호" 에러는 기존 매물 이용
                        is_duplicate = any('이미 등록' in e or 'duplicate' in e.lower() for e in err_list)
                        if is_duplicate:
                            print(f"  [경고] 이미 등록된 차대번호 - 기존 매물에 이미지 업로드 진행")
                        else:
                            print(f"  [오류] 폼 validation 실패:")
                            for e in err_list[:5]:
                                print(f"    - {e}")
                            self._capture_screenshot("form_validation_error")
                            self._last_error_step = '폼_저장_검증'
                            self._last_error_cause = f"validation 오류: {'; '.join(err_list[:3])}"
                            return False

                    # alert 처리
                    await self._async_dismiss_popup()

                    # 24. listing ID 추출 (저장 전후 비교)
                    listing_id = await self._extract_listing_id(
                        car_info.inspection_chassis_no or "",
                        pre_save_ids=pre_save_ids)
                    if not listing_id:
                        print(f"  [경고] listing ID 추출 실패")
                        self._capture_screenshot("listing_id_not_found")
                        self._last_error_step = 'listing_ID_추출'
                        _after_url = self.tab.url if self.tab else ''
                        self._last_error_cause = f"저장 후 URL에서 ID 미발견 (chassis={car_info.inspection_chassis_no}, url={_after_url[:80]})"
                        return False

                    # listing 제출 성공 플래그 (이미지 실패와 무관하게 UPLOADED 기입용)
                    self._listing_submitted = True
                    print(f"  [OK] listing id: {listing_id}")

                    # 25. 이미지 업로드 (버튼 클릭으로 이동)
                    _is_4wd_upload = self._get_drive_type(car_info.car_type or '') == '3'
                    if image_files:
                        print(f"  [INFO] move to photo/upload/{listing_id}")
                        img_ok = await self._async_upload_images_for_listing(listing_id, image_files, is_4wd=_is_4wd_upload)
                        if not img_ok:
                            print(f"  [경고] 이미지 업로드 실패 (매물은 등록됨)")
                            self._capture_screenshot("image_upload_failed")

                    setattr(car_info, '_listing_id', listing_id)

                except Exception as e:
                    print(f"[경고] 저장 중 오류: {e}")
                    self._capture_screenshot("save_error")
                    # ValueError는 모델/제조사 매칭 실패 — 단계/원인 기록
                    if isinstance(e, ValueError) and not self._last_error_step:
                        self._last_error_step = '모델_선택'
                        self._last_error_cause = str(e)
                    elif not self._last_error_step:
                        self._last_error_step = '폼_저장'
                        self._last_error_cause = str(e)
                    import traceback
                    traceback.print_exc()
                    return False
            else:
                pass

            return True

        except Exception as e:
            err_s = str(e).lower()
            print(f"[오류] 차량 정보 입력 중 오류: {e}")
            self._capture_screenshot("fill_vehicle_error")
            import traceback
            traceback.print_exc()
            # 단계/원인이 아직 미기록이면 여기서 캡처
            if not self._last_error_step:
                self._last_error_step = '폼_입력_중_예외'
                self._last_error_cause = str(e)[:200]
            # 탭 크래시/세션 에러는 상위로 재전파 (retry에서 Chrome 재시작 트리거)
            if any(k in err_s for k in ('crashed', 'invalid session', 'disconnected', 'no such window')):
                raise
            return False

    # ── 2단계: 이미지 업로드 (버튼 클릭 기반) ──────────────────────────────────

    async def _async_upload_images_for_listing(self, listing_id: str, image_files: list[str], is_4wd: bool = False) -> bool:
        """photo/upload URL 직접 이동 → 이미지 업로드 → condition-form 처리."""
        try:
            photo_url = f"https://external-vendor.beforward.jp/photo/upload/{listing_id}"
            await self.tab.get(photo_url)
            await asyncio.sleep(2)

            if 'login' in self.tab.url.lower():
                if not self.login():
                    return False
                await self.tab.get(photo_url)
                await asyncio.sleep(2)

            await self._inject_alert_interceptor()
            await self._inject_duplicate_modal_auto_closer()

            # 3. 이미지 업로드
            uploaded = await self._upload_images_to_form(image_files)
            if not uploaded:
                print(f"  [경고] 이미지 업로드 실패")
                self._capture_screenshot("image_form_upload_failed")
                return False

            # 4. 저장 (버튼 클릭 + 60초 대기)
            await self._save_image_upload_page()

            # 5. condition-form 처리
            await self._set_corrosion_no()
            await self._set_condition_div11_button(is_4wd)
            await self._save_after_condition_page()
            return True

        except Exception as e:
            print(f"  [오류] 이미지 업로드 오류: {e}")
            self._capture_screenshot("image_upload_exception")
            import traceback
            traceback.print_exc()
            return False

    # ── JS 기반 Select 헬퍼 ───────────────────────────────────────────────────

    async def _js_select_by_value(self, name: str, value: str, label: str = "") -> bool:
        result = await self.tab.evaluate(f"""
            var sel = document.querySelector('[name="{name}"]');
            if (!sel) return false;
            var opt = Array.from(sel.options).find(o => o.value === {json.dumps(str(value))});
            if (!opt) return false;
            sel.value = {json.dumps(str(value))};
            ['input','change'].forEach(ev => sel.dispatchEvent(new Event(ev, {{bubbles:true}})));
            return true;
        """)
        return bool(result)

    async def _js_select_by_text(self, name: str, text: str, label: str = "", suppress_log: bool = False) -> bool:
        _ = label, suppress_log
        text_lower = text.lower()
        result = await self.tab.evaluate(f"""
            var sel = document.querySelector('[name="{name}"]');
            if (!sel) return null;
            var opt = Array.from(sel.options).find(o => o.text.toLowerCase().includes({json.dumps(text_lower, ensure_ascii=False)}));
            if (!opt) return null;
            sel.value = opt.value;
            ['input','change'].forEach(ev => sel.dispatchEvent(new Event(ev, {{bubbles:true}})));
            return opt.text;
        """)
        return bool(result)

    async def _js_get_select_options(self, name: str) -> list[dict]:
        """select 옵션 목록을 가져와 파싱"""
        result = await self.tab.evaluate(f"""
            var sel = document.querySelector('[name="{name}"]');
            if (!sel || !sel.options) return [];
            return Array.from(sel.options).map(function(o) {{
                return {{text: o.text, value: o.value}};
            }});
        """)
        if not result:
            return []
        if isinstance(result, list):
            return [{'text': str(o.get('text', '')), 'value': str(o.get('value', ''))} for o in result]
        try:
            import json as _json
            parsed = _json.loads(str(result))
            return [{'text': str(o.get('text', '')), 'value': str(o.get('value', ''))} for o in parsed]
        except Exception:
            return []

    async def _css_get_select_options(self, css: str) -> list[dict]:
        """CSS 선택자로 select 요소를 찾아 옵션 목록 반환"""
        result = await self.tab.evaluate(f"""
            var sel = document.querySelector({json.dumps(css)});
            if (!sel || !sel.options) return [];
            return Array.from(sel.options).map(function(o) {{
                return {{text: o.text, value: o.value}};
            }});
        """)
        if not result:
            return []
        if isinstance(result, list):
            return [{'text': str(o.get('text', '')), 'value': str(o.get('value', ''))} for o in result]
        try:
            import json as _json
            parsed = _json.loads(str(result))
            return [{'text': str(o.get('text', '')), 'value': str(o.get('value', ''))} for o in parsed]
        except Exception:
            return []

    async def _css_select_by_value(self, css: str, value: str, label: str = "") -> bool:
        """CSS 선택자로 select 요소를 찾아 value로 선택"""
        result = await self.tab.evaluate(f"""
            var sel = document.querySelector({json.dumps(css)});
            if (!sel) return false;
            var opt = Array.from(sel.options).find(o => o.value === {json.dumps(str(value))});
            if (!opt) return false;
            sel.value = {json.dumps(str(value))};
            ['input','change'].forEach(ev => sel.dispatchEvent(new Event(ev, {{bubbles:true}})));
            return true;
        """)
        return bool(result)

    async def _js_set_select_by_text_exact(self, name: str, text: str, dispatch_events: bool = True) -> bool:
        events_js = "['input','change'].forEach(ev => sel.dispatchEvent(new Event(ev, {bubbles:true})));" if dispatch_events else ""
        result = await self.tab.evaluate(f"""
            var sel = document.querySelector('[name="{name}"]');
            if (!sel) return false;
            var opt = Array.from(sel.options).find(o => o.text === {json.dumps(text, ensure_ascii=False)});
            if (!opt) return false;
            sel.value = opt.value;
            {events_js}
            return true;
        """)
        return bool(result)

    # ── 텍스트 입력 헬퍼 ─────────────────────────────────────────────────────

    async def _fill_text_by_name(self, name: str, value: str, label: str = "") -> bool:  # noqa: ARG002
        if not value:
            return True
        ok = await self.tab.evaluate(f"""
            var el = document.querySelector('[name="{name}"]');
            if (!el) return false;
            el.value = {json.dumps(str(value))};
            ['input','change'].forEach(ev => el.dispatchEvent(new Event(ev, {{bubbles:true}})));
            return true;
        """)
        return bool(ok)

    async def _fill_text_by_id(self, elem_id: str, value: str, label: str = "") -> bool:  # noqa: ARG002
        if not value:
            return True
        ok = await self.tab.evaluate(f"""
            var el = document.getElementById({json.dumps(elem_id)});
            if (!el) return false;
            el.value = {json.dumps(str(value))};
            ['input','change'].forEach(ev => el.dispatchEvent(new Event(ev, {{bubbles:true}})));
            return true;
        """)
        if not ok:
            print(f"  [경고] {label} 필드 없음 (id={elem_id})")
        return bool(ok)

    async def _fill_or_select_by_name(self, name: str, value: str, label: str = "") -> bool:
        """select면 value 선택, input이면 텍스트 입력"""
        result = await self.tab.evaluate(f"""
            var el = document.querySelector('[name="{name}"]');
            if (!el) return false;
            if (el.tagName.toLowerCase() === 'select') {{
                var opt = Array.from(el.options).find(o => o.value === {json.dumps(str(value))});
                if (opt) {{
                    el.value = {json.dumps(str(value))};
                    ['input','change'].forEach(ev => el.dispatchEvent(new Event(ev, {{bubbles:true}})));
                    return true;
                }}
                return false;
            }} else {{
                el.value = {json.dumps(str(value))};
                ['input','change'].forEach(ev => el.dispatchEvent(new Event(ev, {{bubbles:true}})));
                return true;
            }}
        """)
        if not result:
            print(f"  [경고] {label} 입력/선택 실패")
        return bool(result)

    # ── 라디오 버튼 ──────────────────────────────────────────────────────────

    async def _click_radio_by_name(self, name: str, value: str, label: str = "") -> bool:
        ok = await self.tab.evaluate(f"""
            var el = document.querySelector('input[type="radio"][name="{name}"][value="{value}"]');
            if (!el) return false;
            el.scrollIntoView({{block:'center'}});
            el.click();
            el.dispatchEvent(new Event('change', {{bubbles:true}}));
            return true;
        """)
        if not ok:
            print(f"  [경고] {label} 라디오 없음 (name='{name}', value='{value}')")
        return bool(ok)

    # ── XPath 클릭 헬퍼 ──────────────────────────────────────────────────────

    async def _click_xpath(self, xpath: str, label: str = "", timeout: int = 5) -> bool:
        deadline = self._get_loop().time() + timeout
        while True:
            try:
                elems = await self.tab.xpath(xpath)
                if elems:
                    elem = elems[0]
                    await elem.scroll_into_view()
                    await asyncio.sleep(0.1)
                    await elem.click()
                    await asyncio.sleep(0.2)
                    return True
            except Exception:
                pass
            if self._get_loop().time() >= deadline:
                break
            await asyncio.sleep(0.5)
        return False

    async def _click_image_upload_tab(self) -> bool:
        """이미지 업로드 탭 버튼 클릭 (여러 선택자 시도)"""
        # CSS 선택자 우선 (절대 XPath보다 안정적)
        CSS_CANDIDATES = [
            'button.btn-info',
            'input.btn-info',
            'a.btn-info',
        ]
        XPATH_CANDIDATES = [
            '/html/body/div[2]/div[1]/div[1]/div/div/div[1]/div[1]/div[2]/button',
            '//button[contains(@class,"btn-info")]',
            '//a[contains(text(),"画像") or contains(text(),"Image") or contains(text(),"이미지")]',
        ]

        # CSS 시도
        for css in CSS_CANDIDATES:
            try:
                elems = await self.tab.select_all(css)
                if elems:
                    await elems[0].scroll_into_view()
                    await elems[0].click()
                    return True
            except Exception:
                pass

        # XPath 시도
        for xpath in XPATH_CANDIDATES:
            if await self._click_xpath(xpath, "이미지 업로드 탭", timeout=5):
                return True

        print(f"  [경고] 이미지 업로드 탭 버튼을 찾지 못함")
        return False

    async def _get_all_listing_ids(self) -> set:
        """매물 관리 페이지(모든차량 탭)에서 현재 모든 listing ID를 수집한다."""
        await self.tab.get("https://external-vendor.beforward.jp/?limit=50&tab=5")
        await asyncio.sleep(2)
        ids_json = await self.tab.evaluate("""
            (function() {
                var ids = [];
                document.querySelectorAll('a[href]').forEach(function(a) {
                    if (a.href.toLowerCase().indexOf('/tempvehdetails/edit/') !== -1) {
                        var m = a.href.match(/\\/edit\\/(\\d+)/i);
                        if (m) ids.push(m[1]);
                    }
                });
                return JSON.stringify([...new Set(ids)]);
            })()
        """)
        try:
            return set(json.loads(str(ids_json)))
        except Exception:
            return set()

    async def _extract_listing_id(self, chassis_no: str, pre_save_ids: set = None) -> str:
        """저장 후 새로 생성된 listing ID를 추출한다.

        pre_save_ids가 주어지면, 저장 전후 차이(새 ID)를 반환한다.
        """
        try:
            # 방법 0: 현재 URL에 listing ID가 포함되어 있는지 확인
            cur_url = self.tab.url
            import re as _re
            url_match = _re.search(r'/(?:edit|photo/upload)/(\d+)', cur_url)
            if url_match:
                new_id = url_match.group(1)
                if not pre_save_ids or new_id not in pre_save_ids:
                    return new_id

            # 매물 관리 페이지(모든차량 탭)로 이동
            await self.tab.get("https://external-vendor.beforward.jp/?limit=50&tab=5")
            await asyncio.sleep(2)

            # 방법 1: 저장 전 ID 목록과 비교하여 새로 생긴 ID 찾기
            if pre_save_ids:
                new_ids_json = await self.tab.evaluate("""
                    (function() {
                        var ids = [];
                        document.querySelectorAll('a[href]').forEach(function(a) {
                            if (a.href.toLowerCase().indexOf('/tempvehdetails/edit/') !== -1) {
                                var m = a.href.match(/\\/edit\\/(\\d+)/i);
                                if (m) ids.push(m[1]);
                            }
                        });
                        return JSON.stringify([...new Set(ids)]);
                    })()
                """)
                try:
                    current_ids = set(json.loads(str(new_ids_json)))
                except Exception:
                    current_ids = set()

                diff = current_ids - pre_save_ids
                if len(diff) == 1:
                    return diff.pop()
                elif len(diff) > 1:
                    return max(diff, key=int)

            # 방법 2: 차대번호로 테이블 행 전체에서 검색
            listing_id = await self.tab.evaluate(f"""
                (function() {{
                    var chassis = {json.dumps(chassis_no)};
                    if (!chassis) return '';
                    var rows = document.querySelectorAll('table tbody tr, tr');
                    for (var i = 0; i < rows.length; i++) {{
                        var text = rows[i].textContent || '';
                        if (text.indexOf(chassis) !== -1) {{
                            var link = rows[i].querySelector('a[href]');
                            if (link && link.href.toLowerCase().indexOf('/tempvehdetails/edit/') !== -1) {{
                                var m = link.href.match(/\\/edit\\/(\\d+)/i);
                                if (m) return m[1];
                            }}
                        }}
                    }}
                    return '';
                }})()
            """)
            if listing_id:
                return str(listing_id)

            return ""
        except Exception:
            return ""

    async def _click_required_xpaths(self) -> None:
        for xpath in FINAL_CLICK_XPATHS:
            await self._click_xpath(xpath)

    # ── 제조사 선택 ──────────────────────────────────────────────────────────

    async def _select_make(self, make_name: str, car_type: str = "", model_name: str = "") -> bool:
        candidates = [make_name] if make_name else []
        if car_type:
            for kw, make in MODEL_KEYWORD_TO_MAKE.items():
                if kw in car_type and make not in candidates:
                    candidates.append(make)

        alias_map = {
            'MERCEDES BENZ': ['MERCEDES-BENZ', 'BENZ', 'MERCEDES'],
            'RENAULT SAMSUNG': ['RENAULT-SAMSUNG', 'RENAULT', 'SAMSUNG'],
            'CHEVROLET': ['CHEVROLET', 'GM'],
            'LAND ROVER': ['LAND ROVER', 'LANDROVER', 'LAND-ROVER'],
            'VOLKSWAGEN': ['VOLKSWAGEN', 'VW'],
        }
        expanded = list(candidates)
        for c in candidates:
            for alias in alias_map.get(c, []):
                if alias not in expanded:
                    expanded.append(alias)

        return await self._select_make_select2(expanded, model_name)

    async def _select_make_select2(self, candidates: list, model_name: str = '') -> bool:
        """Select2 UI를 통한 제조사 선택 후 즉시 모델 선택"""
        driver = self.tab._driver

        # Select2 컨테이너 클릭해서 드롭다운 열기
        try:
            container = driver.find_element(SeleniumBy.XPATH, '//*[@id="select2-make-id-container"]')
            container.click()
            await asyncio.sleep(0.5)
        except Exception as e:
            print(f"  [경고] 제조사 Select2 컨테이너 클릭 실패: {e}")
            return False

        # 검색창 확인
        try:
            search_input = driver.find_element(SeleniumBy.CSS_SELECTOR, '.select2-search__field')
        except Exception as e:
            print(f"  [경고] 제조사 Select2 검색창 없음: {e}")
            return False

        # 후보 중 첫 번째 단어로 검색해서 가장 유사한 옵션 선택
        search_term = candidates[0].split()[0] if candidates else ""
        search_input.clear()
        search_input.send_keys(search_term)
        await asyncio.sleep(0.5)

        # 현재 보이는 옵션 목록 수집
        try:
            option_elems = driver.find_elements(SeleniumBy.CSS_SELECTOR, '.select2-results__option')
        except Exception:
            option_elems = []

        if not option_elems:
            print(f"  [경고] 제조사 Select2 옵션 없음 (검색어: '{search_term}')")
            # ESC 대신 body 클릭으로 드롭다운 닫기 (ESC가 페이지를 초기화할 수 있음)
            try:
                driver.find_element(SeleniumBy.TAG_NAME, 'body').click()
            except Exception:
                pass
            return False

        # 후보와 가장 유사한 옵션 찾기
        best_elem, best_text, best_score = None, "", 0.0
        for elem in option_elems:
            opt_text = elem.text.strip()
            if not opt_text or opt_text.lower() in ('searching…', 'no results found'):
                continue
            opt_norm = self._normalize_key(opt_text)
            for cand in candidates:
                c_norm = self._normalize_key(cand)
                if c_norm in opt_norm or opt_norm in c_norm:
                    score = 0.95
                else:
                    score = SequenceMatcher(None, c_norm, opt_norm).ratio()
                if score > best_score:
                    best_score, best_text, best_elem = score, opt_text, elem

        if best_elem and best_score >= 0.55:
            best_elem.click()
            await asyncio.sleep(1.0)

            # 제조사 클릭 후 model options AJAX 로드 대기 (최대 5초)
            # ※ change 이벤트를 dispatch하면 2차 AJAX가 폼 전체를 교체하므로 value만 설정 (no event)
            # 모델 선택 + change 이벤트 발생 (모델명 유무 관계없이 항상 실행)
            # → change 이벤트가 2차 AJAX를 트리거해야 연료/미션/색상 select가 채워짐
            import json as _json
            model_set = False
            for _mi in range(10):
                ok = driver.execute_script(f"""
                    var sel = document.querySelector('[name="TempVehDetails[model_id]"]');
                    if (!sel || sel.options.length <= 1) return 'loading';
                    var targetText = {_json.dumps(model_name)};
                    if (!targetText) return 'no_model_name';
                    var opt = Array.from(sel.options).find(o => o.text === targetText);
                    if (!opt) return 'no_exact';
                    if (!opt) return 'no_option';
                    sel.value = opt.value;
                    // change 이벤트 → 2차 AJAX로 연료/미션/색상 등 폼 전체 갱신
                    ['input','change'].forEach(function(ev) {{
                        sel.dispatchEvent(new Event(ev, {{bubbles: true}}));
                    }});
                    return opt.text;
                """)
                if ok in ('no_exact', 'no_model_name'):
                    raise ValueError(f"모델 정확 매칭 없음: '{model_name}' - 엑셀 재원표 확인 필요")
                if ok and ok not in ('loading', 'no_option', False, None):
                    model_set = True
                    break
                await asyncio.sleep(1)

            if not model_set:
                raise ValueError(f"모델 옵션 로드 실패: '{model_name}'")

            # 2차 AJAX 완료 대기: 연료 select에 옵션이 로드될 때까지 (최대 10초)
            _FUEL_CSS = '#bulk_confirm_form div > div > div:first-child table:nth-of-type(3) tr:nth-child(6) select'
            for _wi in range(20):
                await asyncio.sleep(0.5)
                cnt = driver.execute_script(f"""
                    var sel = document.querySelector({json.dumps(_FUEL_CSS)});
                    return sel ? sel.options.length : 0;
                """)
                if cnt and int(cnt) > 1:
                    break

            return True

        print(f"  [경고] 제조사 매칭 실패: {candidates}")
        self._last_error_step = '제조사_선택'
        self._last_error_cause = f"Select2에서 일치하는 제조사 없음 (후보: {candidates})"
        try:
            driver.find_element(SeleniumBy.TAG_NAME, 'body').click()
        except Exception:
            pass
        return False

    # ── 모델 선택 ─────────────────────────────────────────────────────────────

    async def _select_model_fuzzy(self, model_name: str) -> bool:
        if not model_name:
            return False
        # 모델은 일반 <select> 요소 - JS로 직접 처리
        for sel_name in MODEL_SELECT_CANDIDATES:
            options = await self._js_get_select_options(sel_name)
            valid_opts = [o for o in options if o['text'].strip() and o['value']]
            if not valid_opts:
                continue

            model_lower = model_name.lower()
            best_text, best_score = "", 0.0
            for opt in valid_opts:
                opt_lower = opt['text'].lower()
                if model_lower in opt_lower or opt_lower in model_lower:
                    score = 0.95
                else:
                    score = SequenceMatcher(None, model_lower, opt_lower).ratio()
                if score > best_score:
                    best_score, best_text = score, opt['text']

            if best_text and best_score >= 0.55:
                ok = await self._js_set_select_by_text_exact(sel_name, best_text, dispatch_events=True)
                if ok:
                    return True

        print(f"  [경고] 모델 매칭 실패: '{model_name}'")
        return False

    # ── 차량 타입 선택 ────────────────────────────────────────────────────────

    async def _select_body_type(self, body_type: str) -> bool:
        candidates = [body_type]
        for t in BODY_TYPE_TEXT_CANDIDATES.get(body_type, []):
            if t not in candidates:
                candidates.append(t)

        for sel_name in BODY_TYPE_SELECT_CANDIDATES:
            for text in candidates:
                if await self._js_select_by_text(sel_name, text, "차량타입", suppress_log=True):
                    return True

        return False

    # ── 옵션 체크박스 ─────────────────────────────────────────────────────────

    async def _native_click_option(self, input_id: str, label_text: str) -> bool:
        """이미 checked 면 그대로 두고, unchecked 면 Selenium 네이티브 click 으로 체크."""
        try:
            already_checked = await self.tab.evaluate(
                f"var el = document.getElementById({json.dumps(input_id)}); el ? !!el.checked : false;"
            )
            if already_checked:
                return True

            await self.tab.evaluate(
                f"var el = document.getElementById({json.dumps(input_id)}); if (el) el.scrollIntoView({{block:'center'}});"
            )

            from selenium.webdriver.common.by import By as _By
            elem = self.browser.find_element(_By.ID, input_id)
            try:
                elem.click()
            except Exception:
                try:
                    self.browser.execute_script(
                        "var lbl = document.querySelector('label[for=\"' + arguments[0] + '\"]');"
                        "if (lbl) lbl.click();", input_id
                    )
                except Exception:
                    return False

            checked = await self.tab.evaluate(
                f"var el = document.getElementById({json.dumps(input_id)}); el ? el.checked : false;"
            )
            return bool(checked)
        except Exception as e:
            print(f"  [경고] {label_text} 네이티브 click 실패: {e}")
            return False

    async def _fill_options(self, options) -> None:
        checked_count = 0
        for opt in options:
            if not opt.mapped_name:
                continue
            # 라벨 매칭으로 input ID 찾기 (공백/특수문자 무시한 부분일치)
            input_id = await self.tab.evaluate(f"""
                function norm(s) {{ return (s || '').toLowerCase().replace(/[\\s\\-_·／]+/g, ''); }}
                var name = norm({json.dumps(opt.mapped_name)});
                var labels = document.querySelectorAll('label');
                for (var i = 0; i < labels.length; i++) {{
                    if (norm(labels[i].textContent).indexOf(name) === -1) continue;
                    var forId = labels[i].getAttribute('for');
                    if (forId) {{
                        var t = document.getElementById(forId);
                        if (t && (t.type === 'checkbox' || t.type === 'radio')) return forId;
                    }}
                    var inner = labels[i].querySelector('input[type="checkbox"], input[type="radio"]');
                    if (inner) {{
                        if (!inner.id) inner.id = '_opt_' + Math.random().toString(36).substring(2);
                        return inner.id;
                    }}
                }}
                return '';
            """)

            if not input_id:
                print(f"  [경고] 옵션 라벨 못찾음: {opt.mapped_name}")
                continue

            # Selenium 네이티브 click(트rusted event) 으로 BeForward 인터랙션 추적 우회
            ok = await self._native_click_option(input_id, opt.mapped_name)
            if ok:
                checked_count += 1

    # ── 이미지 업로드 ─────────────────────────────────────────────────────────

    async def _upload_images_to_form(self, image_files: list[str]) -> bool:
        """input[type=file]에 send_keys로 직접 경로 전달 (파일 다이얼로그 우회)"""
        try:
            if not image_files:
                return False

            abs_paths = [os.path.abspath(p) for p in image_files if os.path.exists(p)]
            if not abs_paths:
                return False

            # 기존 이미지 삭제
            await self._delete_existing_images()

            # input[type=file] 찾기 (숨겨진 상태이므로 JS로 visible 처리)
            file_inputs = self.browser.find_elements(SeleniumBy.CSS_SELECTOR,
                '#public_pane input[type="file"]')
            if not file_inputs:
                file_inputs = self.browser.find_elements(SeleniumBy.CSS_SELECTOR,
                    'input[type="file"]')
            if not file_inputs:
                return False

            file_input = file_inputs[0]

            # ChromeDriver는 hidden input에도 send_keys 가능 - JS visibility 조작 불필요
            # 여러 파일은 \n 구분으로 한 번에 전달 (multiple 속성 있을 때)
            paths_str = "\n".join(abs_paths)
            file_input.send_keys(paths_str)
            await asyncio.sleep(1)
            return True

        except Exception as e:
            print(f"  [경고] 이미지 업로드 실패: {e}")
            import traceback; traceback.print_exc()
            try:
                pyautogui.press('escape')
            except Exception:
                pass
            return False

    async def _wait_image_save_complete(self, timeout: int = 300) -> None:
        """이미지 저장 버튼 클릭 후, 서버 처리 완료까지 대기한다.

        progress bar, spinner, 로딩 표시가 모두 사라질 때까지 폴링.
        최소 10초, 최대 timeout초 대기.
        """
        await asyncio.sleep(5)
        deadline = self._get_loop().time() + timeout
        stable_count = 0
        while self._get_loop().time() < deadline:
            is_busy = await self.tab.evaluate("""
                (function() {
                    var sels = [
                        '.progress:not([style*="none"])',
                        '.uploading', '.loading', '.spinner',
                        '[class*="progress"]:not([style*="none"])',
                        '[class*="loading"]', '[class*="spinner"]',
                        '.modal.in', '.modal.show',
                        '.blockUI',
                    ];
                    for (var i = 0; i < sels.length; i++) {
                        var els = document.querySelectorAll(sels[i]);
                        for (var j = 0; j < els.length; j++) {
                            if (els[j].offsetParent !== null) return true;
                        }
                    }
                    return false;
                })()
            """)
            if not is_busy:
                stable_count += 1
                if stable_count >= 3:
                    return
            else:
                stable_count = 0
            await asyncio.sleep(2)
        print(f"  [경고] 이미지 저장 대기 타임아웃 ({timeout}초)")

    async def _find_file_input(self):
        """file input 요소 찾기.
        //*[@id="public_pane"]/ul/li/label 이 파일탐색기를 여는 라벨이므로
        해당 라벨 하위의 input[type="file"] 을 우선 탐색.
        """
        # input을 숨김 해제
        try:
            await self.tab.evaluate("""
                document.querySelectorAll('input[type="file"]').forEach(function(inp) {
                    inp.style.display = 'block';
                    inp.style.visibility = 'visible';
                    inp.style.opacity = '0.01';
                    inp.style.position = 'absolute';
                });
            """)
        except Exception:
            pass

        # 1순위: 사용자가 확인한 라벨 내부 input
        for xpath in [
            '//*[@id="public_pane"]/ul/li/label//input[@type="file"]',
            '//*[@id="public_pane"]/ul/li/label/input[@type="file"]',
            '//*[@id="public_pane"]/ul/li//input[@type="file"]',
        ]:
            try:
                elems = await self.tab.xpath(xpath)
                if elems:
                    return elems[0]
            except Exception:
                pass

        # 2순위: public_pane 전체에서 탐색
        try:
            inputs = await self.tab.select_all('#public_pane input[type="file"]')
            if inputs:
                return inputs[0]
        except Exception:
            pass

        # 3순위: 페이지 전체
        try:
            inputs = await self.tab.select_all('input[type="file"]')
            if inputs:
                return inputs[0]
        except Exception:
            pass

        return None

    async def _delete_existing_images(self) -> int:
        """기존 이미지를 1개씩 삭제. 실제 페이지 구조를 먼저 진단한다."""

        # 1. 페이지에 이미 업로드된 이미지 수 확인
        existing_count = await self.tab.evaluate("""
            (function() {
                var pane = document.querySelector('#public_pane, #image_pane, .photo-list, .image-list');
                if (!pane) return -1;
                var imgs = pane.querySelectorAll('img, .photo-item, .image-item, li.uploaded, li[data-id]');
                return imgs.length;
            })()
        """)
        existing_count = existing_count or 0
        if existing_count == 0:
            return 0

        # 2. 실제 삭제 버튼 탐색 (페이지 구조 진단 포함)
        page_info = await self.tab.evaluate("""
            (function() {
                var result = {found: [], count: 0};
                // 삭제 가능한 모든 버튼/링크 탐색
                var all = document.querySelectorAll('a, button, span, i');
                all.forEach(function(el) {
                    if (el.offsetParent === null) return;
                    var cls = (el.className || '').toLowerCase();
                    var txt = (el.textContent || '').trim().toLowerCase();
                    var aria = (el.getAttribute('aria-label') || '').toLowerCase();
                    var title = (el.getAttribute('title') || '').toLowerCase();
                    if (cls.includes('delete') || cls.includes('remove') || cls.includes('delet') ||
                        txt === 'delete' || txt === 'remove' || txt === '×' || txt === 'x' ||
                        aria.includes('delete') || aria.includes('remove') ||
                        title.includes('delete') || title.includes('remove')) {
                        result.found.push({tag: el.tagName, cls: el.className, txt: txt.slice(0,20)});
                        result.count++;
                    }
                });
                return result;
            })()
        """)
        if page_info and page_info.get('count', 0) == 0 and existing_count != 0:
            self._capture_screenshot("delete_btn_not_found")

        # 3. 삭제 실행 (1개씩)
        deleted = 0
        for _round in range(50):
            btn_found = await self.tab.evaluate("""
                (function() {
                    var all = document.querySelectorAll('a, button, span, i, input[type="button"]');
                    for (var i = 0; i < all.length; i++) {
                        var el = all[i];
                        if (el.offsetParent === null) continue;
                        var cls = (el.className || '').toLowerCase();
                        var txt = (el.textContent || '').trim().toLowerCase();
                        var aria = (el.getAttribute('aria-label') || '').toLowerCase();
                        var title = (el.getAttribute('title') || '').toLowerCase();
                        if (cls.includes('delete') || cls.includes('remove') ||
                            aria.includes('delete') || aria.includes('remove') ||
                            title.includes('delete') || title.includes('remove')) {
                            el.click();
                            return el.className + '|' + txt.slice(0,20);
                        }
                    }
                    return false;
                })()
            """)
            if not btn_found:
                break

            await asyncio.sleep(0.5)

            # 확인 팝업 (OK / Yes 등) 처리
            await self.tab.evaluate("""
                (function() {
                    var btns = document.querySelectorAll('button, a, input[type="button"]');
                    for (var i = 0; i < btns.length; i++) {
                        if (btns[i].offsetParent === null) continue;
                        var t = btns[i].textContent.trim().toUpperCase();
                        if (t === 'OK' || t === 'YES' || t === 'CONFIRM' || t === '확인') {
                            btns[i].click(); return true;
                        }
                    }
                    // lightbox 확인
                    var lb = document.querySelector('#lightbox a, #lightbox button');
                    if (lb && lb.offsetParent !== null) { lb.click(); return true; }
                    return false;
                })()
            """)
            try:
                await self.tab.evaluate("window._pendingAlert = ''")
            except Exception:
                pass

            await asyncio.sleep(1)
            deleted += 1

        if deleted > 0:
            await asyncio.sleep(1)
        return deleted

    async def _wait_upload_complete(self, timeout: int = 60) -> None:
        """업로드/처리 다이얼로그('이미지 파일을 업로드 중입니다' 포함)가 사라질 때까지 대기"""
        await asyncio.sleep(2)
        deadline = self._get_loop().time() + timeout
        _screenshot_taken = False
        while self._get_loop().time() < deadline:
            try:
                result = await self.tab.evaluate("""
                    (function() {
                        // 진행바/스피너 체크
                        var elems = document.querySelectorAll(
                            '.progress:not([style*="none"]), .uploading, .loading, [class*="progress"]:not([style*="none"])');
                        if (Array.from(elems).some(function(el) { return el.offsetParent !== null; }))
                            return 'busy';
                        // '업로드 중' 텍스트 다이얼로그 체크
                        var body = document.body.innerText || '';
                        if (body.includes('업로드 중') || body.includes('uploading') ||
                            body.includes('処理中') || body.includes('処理')) {
                            // 진행 상태 숫자 추출 (e.g. "12 / 12")
                            var m = body.match(/(\\d+)\\s*\\/\\s*(\\d+)/);
                            if (m && m[1] === m[2]) return 'complete_' + m[1];
                            return 'busy';
                        }
                        return 'done';
                    })()
                """)
                if result and result.startswith('complete_') and not _screenshot_taken:
                    _screenshot_taken = True
                    self._capture_screenshot("image_upload_complete")
                if result == 'done':
                    return
            except Exception:
                return
            await asyncio.sleep(1)

    async def _close_duplicate_modal(self) -> bool:
        try:
            # JS alert 체크
            alert_text = await self.tab.evaluate("window._pendingAlert || ''")
            if alert_text:
                await self.tab.evaluate("window._pendingAlert = ''")
                return True
        except Exception:
            pass

        try:
            closed = await self.tab.evaluate("""
                (function() {
                    // 페이지 내 모든 visible 버튼/링크를 순회하며 닫기 시도
                    var candidates = Array.from(document.querySelectorAll('button, a, input[type="button"], input[type="submit"]'));
                    // 1. OK 텍스트 버튼
                    for (var i = 0; i < candidates.length; i++) {
                        var el = candidates[i];
                        if (el.offsetParent === null) continue;
                        var t = el.textContent.trim().toUpperCase();
                        var v = (el.value || '').trim().toUpperCase();
                        if (t === 'OK' || v === 'OK') { el.click(); return 'ok-btn'; }
                    }
                    // 2. .close / data-dismiss 버튼
                    for (var i = 0; i < candidates.length; i++) {
                        var el = candidates[i];
                        if (el.offsetParent === null) continue;
                        if (el.classList.contains('close') || el.getAttribute('data-dismiss') === 'modal') {
                            el.click(); return 'close-btn';
                        }
                    }
                    // 3. 경고/모달 레이어 내 아무 버튼
                    var layers = document.querySelectorAll('.modal, .dialog, [role="dialog"], #lightbox, .alert-box, .warning');
                    for (var i = 0; i < layers.length; i++) {
                        if (layers[i].offsetParent === null) continue;
                        var inner = layers[i].querySelectorAll('button, a.btn, input[type="button"]');
                        for (var j = 0; j < inner.length; j++) {
                            if (inner[j].offsetParent !== null) { inner[j].click(); return 'layer-btn'; }
                        }
                    }
                    return false;
                })()
            """)
            if closed:
                await asyncio.sleep(0.3)
                return True
            return False
        except Exception:
            return False

    async def _wait_per_file_upload(self, max_seconds: int = 15) -> None:
        """파일 1개 업로드 후 서버 처리 완료까지 대기.
        '업로드 중' 다이얼로그가 나타났다가 사라지면 완료.
        나타나지 않으면 max_seconds/3 후 진행.
        """
        deadline = self._get_loop().time() + max_seconds
        appeared = False

        # Phase 1: 다이얼로그 등장 대기 (최대 3초)
        for _ in range(6):
            busy = await self.tab.evaluate(
                "document.body.innerText.includes('업로드 중') || "
                "document.body.innerText.includes('uploading')"
            )
            if busy:
                appeared = True
                break
            await asyncio.sleep(0.5)

        if not appeared:
            # 다이얼로그 없이 바로 완료 → 0.5s 안전 대기
            await asyncio.sleep(0.5)
            return

        # Phase 2: 다이얼로그 사라질 때까지 대기
        while self._get_loop().time() < deadline:
            busy = await self.tab.evaluate(
                "document.body.innerText.includes('업로드 중') || "
                "document.body.innerText.includes('uploading')"
            )
            if not busy:
                return
            await asyncio.sleep(0.5)


    # ── 이미지 저장 / 성능 페이지 이동 ───────────────────────────────────────

    async def _click_option_radio(self, row: int) -> bool:
        """table[5] 특정 행의 라디오 버튼을 label 클릭 또는 JS로 선택."""
        BASE = '//*[@id="bulk_confirm_form"]/div/div/div[2]/table[5]/tbody'
        label_xpath = f'{BASE}/tr[{row}]/td[3]/label'
        ok = await self._click_xpath(label_xpath, f"옵션 tr[{row}] label", timeout=5)
        if not ok:
            # JS fallback: label 또는 연결된 radio input 직접 클릭
            ok = await self.tab.evaluate(f"""
                (function() {{
                    var rows = document.querySelectorAll('#bulk_confirm_form div > div > div:nth-child(2) table:nth-of-type(5) tbody tr');
                    var tr = rows[{row - 1}];
                    if (!tr) return false;
                    var lbl = tr.querySelector('td:nth-child(3) label');
                    if (lbl) {{ lbl.click(); return true; }}
                    var inp = tr.querySelector('input[type="radio"]');
                    if (inp) {{ inp.checked = true; inp.dispatchEvent(new Event('change', {{bubbles:true}})); return true; }}
                    return false;
                }})()
            """)
            if not ok:
                print(f"  [경고] 옵션 tr[{row}] 라디오 선택 실패")
        return bool(ok)

    async def _click_photo_page_options(self, is_4wd: bool = False) -> None:
        """차량 정보 입력 시 옵션 라디오 선택.
        - tr[2] : 항상 클릭
        - tr[8] : 4WD 차량만 클릭
        """
        await self._click_option_radio(2)

        if is_4wd:
            await self._click_option_radio(8)

    async def _save_image_upload_page(self) -> bool:
        SAVE_XPATH = '//*[@id="bulk_confirm_form"]/div/button'

        # 업로드 썸네일 대기 (최대 30초)
        prev_count = 0
        for tick in range(30):
            try:
                count = self.browser.execute_script("""
                    (function() {
                        var pane = document.querySelector('#public_pane');
                        if (!pane) return 0;
                        var items = pane.querySelectorAll('li img, li.uploaded, li[data-id]');
                        return items.length;
                    })()
                """) or 0
            except Exception:
                count = 0
            if count != prev_count:
                prev_count = count
            if count > 0 and tick >= 2:
                break
            await asyncio.sleep(1)

        try:
            btn = self.browser.find_element(SeleniumBy.XPATH, SAVE_XPATH)
        except Exception:
            return False

        # scrollIntoView 후 ActionChains 클릭
        self.browser.execute_script("arguments[0].scrollIntoView({block:'center'})", btn)
        await asyncio.sleep(1)

        actions = ActionChains(self.browser)
        actions.move_to_element(btn).click().perform()

        # 저장 완료 감지: URL이 변경되거나 저장 버튼이 사라지면 완료 (최대 60초)
        pre_url = self.tab.url
        for _tick in range(60):
            await asyncio.sleep(1)
            try:
                cur_url = self.tab.url
                if cur_url != pre_url:
                    break
                # 저장 버튼이 사라졌는지 확인
                remaining = self.browser.find_elements(SeleniumBy.XPATH, SAVE_XPATH)
                if not remaining:
                    break
            except Exception:
                break
        return True

    async def _pyautogui_click_xpath(self, xpath: str, label: str = "") -> bool:
        """pyautogui로 OS 레벨 실제 마우스 클릭 (manual click과 동일)."""
        try:
            import pyautogui
            pyautogui.FAILSAFE = False

            # 버튼 위치 계산 (브라우저 window + chrome 높이 + 요소 위치)
            pos = await self.tab.evaluate(f"""
                (function() {{
                    var btn = document.evaluate({json.dumps(xpath)}, document, null,
                        XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue;
                    if (!btn) return null;
                    btn.scrollIntoView({{block: 'center'}});
                    // scrollIntoView 후 위치 재계산을 위해 rAF 없이 즉시 읽기
                    var r = btn.getBoundingClientRect();
                    var chromeH = window.outerHeight - window.innerHeight;
                    return {{
                        x: Math.round(window.screenX + r.left + r.width / 2),
                        y: Math.round(window.screenY + chromeH + r.top + r.height / 2),
                        dpr: window.devicePixelRatio || 1,
                        w: Math.round(r.width),
                        h: Math.round(r.height),
                        visible: r.width > 0 && r.height > 0 && btn.offsetParent !== null
                    }};
                }})()
            """)
            if not pos:
                return False
            if not pos.get('visible'):
                return False

            screen_x = pos['x']
            screen_y = pos['y']

            await asyncio.sleep(0.3)
            pyautogui.moveTo(screen_x, screen_y, duration=0.3)
            await asyncio.sleep(0.2)
            pyautogui.click()
            return True

        except ImportError:
            return False
        except Exception as e:
            return False

    async def _go_to_condition_page_after_image_upload(self) -> bool:
        """이미지 업로드 후 성능 페이지 이동 버튼 클릭"""
        target_xpath = '//*[@id="bulk_confirm_form"]/div/a[2]'
        try:
            ok = await self._click_xpath(target_xpath, "성능 페이지 이동 버튼", timeout=5)
            return ok
        except Exception as e:
            return False

    async def _set_corrosion_no(self) -> bool:
        xpath = '//*[@id="condition-form"]/div[1]/div/div[1]/div/button[1]'
        ok = await self._click_xpath(xpath, "부식 NO", timeout=5)
        if ok:
            await asyncio.sleep(0.5)
        return ok

    async def _set_condition_div11_button(self, is_4wd: bool) -> bool:
        xpath = ('//*[@id="condition-form"]/div[1]/div/div[11]/div/button[1]' if is_4wd
                 else '//*[@id="condition-form"]/div[1]/div/div[11]/div/button[3]')
        label = "4륜" if is_4wd else "없음"
        ok = await self._click_xpath(xpath, label, timeout=5)
        if ok:
            await asyncio.sleep(0.5)
        return ok

    async def _save_after_condition_page(self) -> bool:
        SAVE_XPATHS = [
            '//*[@id="condition-form"]/div[2]/div/button',
            '//*[@id="bulk_confirm_form"]/div/button',
            SUBMIT_BUTTON_XPATH,
            "//button[@type='submit']",
            "//input[@type='submit']",
        ]
        for xpath in SAVE_XPATHS:
            result = await self.tab.evaluate(f"""
                (function() {{
                    var btn = document.evaluate({json.dumps(xpath)}, document, null,
                        XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue;
                    if (!btn || btn.offsetParent === null) return false;
                    btn.removeAttribute('disabled');
                    btn.scrollIntoView({{block:'center'}});
                    btn.click();
                    return true;
                }})()
            """)
            if result:
                await asyncio.sleep(1.5)
                await self._async_dismiss_popup()
                return True
        return False

    async def _async_dismiss_popup(self) -> str:
        """팝업/alert 닫기. alert 텍스트를 반환."""
        POPUP_CLOSE_XPATH = '//*[@id="lightbox"]/div[2]/div/div[2]/a[1]'

        # 1. JS alert (interceptor로 저장된 것)
        try:
            alert_text = await self.tab.evaluate("window._pendingAlert || ''")
            if alert_text:
                await self.tab.evaluate("window._pendingAlert = ''")
                return str(alert_text)
        except Exception:
            pass

        # 2. lightbox 팝업
        try:
            btns = await self.tab.xpath(POPUP_CLOSE_XPATH)
            if btns:
                await btns[0].click()
                await asyncio.sleep(0.3)
                return ""
        except Exception:
            pass

        # 3. 일반 close 버튼
        try:
            closed = await self.tab.evaluate("""
                (function() {
                    var btns = document.querySelectorAll('.close, [aria-label="Close"], [data-dismiss="modal"]');
                    for (var i = 0; i < btns.length; i++) {
                        if (btns[i].offsetParent !== null) { btns[i].click(); return true; }
                    }
                    return false;
                })()
            """)
            if closed:
                await asyncio.sleep(0.3)
        except Exception:
            pass

        return ""

    # ── 도우미 메서드 ─────────────────────────────────────────────────────────

    def _normalize_key(self, text: str) -> str:
        return re.sub(r'[^a-z0-9]', '', (text or "").lower())

    def _get_make_from_vin(self, vin: str) -> str:
        if not vin or len(vin) < 3:
            return ""
        return WMI_TO_MAKE.get(vin[:3].upper(), "")

    def _get_make_name(self, car_type: str) -> str:
        if not car_type:
            return ""
        text = car_type.strip()
        first_word = text.split()[0] if text.split() else ""
        if first_word in MAKE_MAP:
            return MAKE_MAP[first_word]
        for kw, make in MODEL_KEYWORD_TO_MAKE.items():
            if kw in text:
                return make
        return MAKE_MAP.get(first_word, first_word)

    def _get_body_type(self, car_type: str) -> str:
        if not car_type:
            return 'Sedan'
        car_upper = car_type.upper()
        for body, keywords in BODY_TYPE_MAP.items():
            for kw in keywords:
                if kw.upper() in car_upper:
                    return body
        return 'Sedan'

    def _get_drive_type(self, car_type: str) -> str:
        if not car_type:
            return '2'
        upper = car_type.upper()
        for kw in self._4WD_KEYWORDS_EN:
            if kw in upper:
                return '3'
        for kw in self._4WD_KEYWORDS_KO:
            if kw in car_type:
                return '3'
        return '2'

    def _map_color(self, color: str) -> str:
        if not color:
            return color
        c = color.strip()
        # 1. 정확 일치
        mapped = COLOR_MAP_KO.get(c) or COLOR_MAP_KO.get(c.lower())
        if mapped:
            return mapped
        # 2. 공백 제거 후 부분 문자열 매칭
        color_ns = c.replace(' ', '').lower()
        for k, v in COLOR_MAP_KO.items():
            k_ns = k.replace(' ', '').lower()
            if k_ns in color_ns or color_ns in k_ns:
                return v
        return color

    def _map_fuel(self, fuel_type: str) -> str:
        if not fuel_type:
            return ""
        ft = fuel_type.strip()
        # 1. 정확 일치
        if ft in FUEL_MAP:
            return FUEL_MAP[ft]
        if ft.lower() in FUEL_MAP:
            return FUEL_MAP[ft.lower()]
        # 2. 부분 포함 (가장 긴 키 우선)
        for k in sorted(FUEL_MAP, key=len, reverse=True):
            if k.lower() in ft.lower():
                return FUEL_MAP[k]
        return ft

    def _map_transmission(self, transmission: str) -> str:
        if not transmission:
            return ""
        for k, v in TRANSMISSION_MAP.items():
            if k.lower() in transmission.lower():
                return v
        return transmission

    def _download_images_from_drive_link(self, drive_link: str, row_num: int) -> list[str]:
        if not drive_link:
            return []
        try:
            from drive_image_downloader import DriveImageDownloader
            downloader = DriveImageDownloader()
            downloader.setup_driver()
            try:
                files = downloader.download_images(drive_link, row_num or 0)
                return files
            finally:
                downloader.close()
        except Exception as e:
            print(f"  [경고] 드라이브 이미지 다운로드 실패: {e}")
            import traceback
            traceback.print_exc()
            return []

    def _download_images_from_mango_link(self, mango_url: str, row_num: int) -> list[str]:
        if not mango_url:
            return []
        try:
            from drive_image_downloader import MangocarImageDownloader
            downloader = MangocarImageDownloader()
            downloader.setup_driver()
            try:
                files = downloader.download_images(mango_url, row_num or 0)
                return files
            finally:
                downloader.close()
        except Exception as e:
            print(f"  [경고] 망고카 이미지 다운로드 실패: {e}")
            import traceback
            traceback.print_exc()
            return []

    # ── Google Sheets 모델 레퍼런스 (사용 시에만 로드) ────────────────────────

    def _load_make_model_reference(self) -> None:
        if self._bf_model_ref_loaded:
            return
        self._bf_model_ref_loaded = True
        try:
            import gspread
            from google.oauth2.service_account import Credentials
            from config import SERVICE_ACCOUNT_FILE
            scopes = [
                'https://www.googleapis.com/auth/spreadsheets.readonly',
                'https://www.googleapis.com/auth/drive.readonly',
            ]
            creds = Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=scopes)
            gc = gspread.authorize(creds)
            ss = gc.open_by_key(BF_MAKE_MODEL_SHEET_ID)
            ws = None
            for title in BF_MAKE_MODEL_WORKSHEETS:
                try:
                    ws = ss.worksheet(title)
                    break
                except Exception:
                    continue
            if ws is None:
                ws = ss.get_worksheet(0)
            if ws is None:
                return
            rows = ws.get_all_values()
            if len(rows) < 2:
                return
            headers = [h.strip().lower() for h in rows[0]]
            make_idx, model_idx = 0, 1
            for i, h in enumerate(headers):
                if h in ('메이커', 'maker', 'manufacturer'):
                    make_idx = i
                if h in ('모델', 'model'):
                    model_idx = i
            make_models = {}
            for row in rows[1:]:
                if len(row) <= max(make_idx, model_idx):
                    continue
                make = row[make_idx].strip().upper()
                model = row[model_idx].strip()
                if make and model:
                    make_models.setdefault(make, set()).add(model)
            self._bf_make_models = {k: sorted(v) for k, v in make_models.items()}
        except Exception as e:
            pass
